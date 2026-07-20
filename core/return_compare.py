# -*- coding: utf-8 -*-
"""退貨物流核對工具：客戶退貨授權明細 × HCT 退貨入庫格式。

比對 NetSuite「客戶退貨授權明細報表」（RA，客戶申請退貨的授權明細）與
倉儲端「HCT 退貨入庫格式」（實際入庫收貨記錄），確認退貨數量是否一致。
兩份檔案順序不限，程式依欄位自動辨識類型。

兩邊都沒有可以互相對應的單號欄位（RA 是 RA-DW-xxx 的授權單號，入庫格式
是倉儲自訂的收貨批次代號，兩者無法對應），所以用「料號＋效期」加總數量
比較，比對邏輯與 compare.py 的數量核對一致：
  - RA 明細：DR_料號 ＋ 交易序號/批號（此欄位本身即效期字串）；運費等
    非品項明細列 DR_料號空白，予以排除。
  - HCT 入庫：產品編號（開頭常見「DR」前綴，需與 RA 的 DR_料號對齊）
    ＋ 效期。
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime

from .xlio import normalize_number, normalize_text, read_workbook, try_parse_date, first_sheet

SOURCE_RA = "RETURN_AUTH"
SOURCE_HCT = "HCT_RECEIVE"
SOURCE_AMBIGUOUS = "AMBIGUOUS"

_RA_HEADERS = ["文件編號", "客戶", "DR_料號", "交易序號/批號", "交易序號/批號數量", "倉別"]
_HCT_HEADERS = ["退貨單號", "產品編號", "效期", "數量", "儲區類別"]

STATUS_COLORS = {
    "一致": "E2EFDA",
    "數量不符": "FFEB9C",
    "僅退貨授權": "F4CCCC",
    "僅入庫記錄": "BDD7EE",
}


class ReturnCompareError(ValueError):
    pass


@dataclass
class ReturnCompareResult:
    output_bytes: bytes
    output_name: str
    total_rows: int
    status_counts: dict


def compare(data1: bytes, name1: str, data2: bytes, name2: str) -> ReturnCompareResult:
    rows1 = first_sheet(read_workbook(data1, name1))
    rows2 = first_sheet(read_workbook(data2, name2))
    headers1 = _build_header_map(rows1[0]) if rows1 else {}
    headers2 = _build_header_map(rows2[0]) if rows2 else {}

    type1 = _detect_type(headers1)
    type2 = _detect_type(headers2)
    _validate_types(type1, type2, name1, name2)

    if type1 == SOURCE_RA:
        ra_rows, ra_headers = rows1, headers1
        hct_rows, hct_headers = rows2, headers2
    else:
        ra_rows, ra_headers = rows2, headers2
        hct_rows, hct_headers = rows1, headers1

    detail_rows = _build_compare(ra_rows, ra_headers, hct_rows, hct_headers)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        output_bytes = _write_output(name1, name2, detail_rows)
    except Exception as exc:
        raise ReturnCompareError(f"建立輸出檔時發生錯誤：{exc}") from exc

    return ReturnCompareResult(
        output_bytes=output_bytes,
        output_name=f"退貨物流核對結果_{stamp}.xlsx",
        total_rows=len(detail_rows),
        status_counts=_count_statuses(detail_rows),
    )


# ------------------------------------------------------------------ 讀表輔助


def _build_header_map(header_row: list) -> dict[str, int]:
    headers: dict[str, int] = {}
    for index, value in enumerate(header_row):
        text = normalize_text(value)
        if text and text.casefold() not in headers:
            headers[text.casefold()] = index
    return headers


def _cell(rows: list, row_index: int, headers: dict, name: str) -> object:
    index = headers.get(name.casefold())
    if index is None:
        return None
    row = rows[row_index]
    return row[index] if index < len(row) else None


def _detect_type(headers: dict) -> str:
    has_ra = all(h.casefold() in headers for h in _RA_HEADERS)
    has_hct = all(h.casefold() in headers for h in _HCT_HEADERS)
    if has_ra and has_hct:
        return SOURCE_AMBIGUOUS
    if has_ra:
        return SOURCE_RA
    if has_hct:
        return SOURCE_HCT
    return ""


def _validate_types(type1: str, type2: str, name1: str, name2: str) -> None:
    if SOURCE_AMBIGUOUS in (type1, type2):
        raise ReturnCompareError("有檔案同時符合兩種表格欄位，無法自動判斷。")
    if not type1 or not type2:
        raise ReturnCompareError(
            "無法辨識其中一個檔案的欄位。\n"
            f"檔案一：{name1}，類型：{type1 or '無法辨識'}\n"
            f"檔案二：{name2}，類型：{type2 or '無法辨識'}\n\n"
            "客戶退貨授權明細需要欄位：" + "、".join(_RA_HEADERS) + "\n"
            "HCT 退貨入庫格式需要欄位：" + "、".join(_HCT_HEADERS)
        )
    if type1 == type2:
        raise ReturnCompareError("請選擇一份客戶退貨授權明細與一份 HCT 退貨入庫格式，不要選到同類型檔案。")


# ------------------------------------------------------------------ 正規化


def _strip_dr_prefix(text: str) -> str:
    """HCT 入庫格式的產品編號常是「DR」+ 料號；RA 明細的 DR_料號沒有這個前綴。"""
    if len(text) > 2 and text[:2].casefold() == "dr" and text[2].isdigit():
        return text[2:]
    return text


def _normalize_expiry(value: object) -> tuple[str, str]:
    """回傳 (比對用 key, 顯示用文字)；無法辨識為日期時 key 留空、顯示原始文字。"""
    text = normalize_text(value)
    if not text:
        return "", ""
    parsed = try_parse_date(value)
    if parsed:
        display = parsed.strftime("%Y-%m-%d")
        return display, display
    return "", text


# ------------------------------------------------------------------ 核對


def _build_compare(ra_rows, ra_headers, hct_rows, hct_headers) -> list[dict]:
    records: dict[tuple, dict] = {}

    def get_record(material: str, expiry_key: str, expiry_display: str, name: str) -> dict:
        key = (material.casefold(), expiry_key)
        record = records.setdefault(key, {
            "料號": material, "品名": name, "效期": expiry_display,
            "退貨授權數量": 0.0, "入庫數量": 0.0,
        })
        if not record["品名"] and name:
            record["品名"] = name
        return record

    for row_index in range(1, len(ra_rows)):
        material = _strip_dr_prefix(normalize_text(_cell(ra_rows, row_index, ra_headers, "DR_料號")))
        if not material:
            continue  # 運費等非品項明細列，DR_料號空白
        expiry_key, expiry_display = _normalize_expiry(_cell(ra_rows, row_index, ra_headers, "交易序號/批號"))
        quantity = normalize_number(_cell(ra_rows, row_index, ra_headers, "交易序號/批號數量"))
        name = normalize_text(_cell(ra_rows, row_index, ra_headers, "顯示名稱"))
        record = get_record(material, expiry_key, expiry_display, name)
        record["退貨授權數量"] += quantity

    for row_index in range(1, len(hct_rows)):
        material = _strip_dr_prefix(normalize_text(_cell(hct_rows, row_index, hct_headers, "產品編號")))
        if not material:
            continue
        expiry_key, expiry_display = _normalize_expiry(_cell(hct_rows, row_index, hct_headers, "效期"))
        quantity = normalize_number(_cell(hct_rows, row_index, hct_headers, "數量"))
        name = normalize_text(_cell(hct_rows, row_index, hct_headers, "產品名稱"))
        record = get_record(material, expiry_key, expiry_display, name)
        record["入庫數量"] += quantity

    result = []
    for record in records.values():
        ra_qty, hct_qty = record["退貨授權數量"], record["入庫數量"]
        record["差異(授權-入庫)"] = ra_qty - hct_qty
        if ra_qty == 0 and hct_qty != 0:
            record["狀態"] = "僅入庫記錄"
        elif ra_qty != 0 and hct_qty == 0:
            record["狀態"] = "僅退貨授權"
        elif abs(ra_qty - hct_qty) < 1e-6:
            record["狀態"] = "一致"
        else:
            record["狀態"] = "數量不符"
        result.append(record)
    result.sort(key=lambda r: (r["料號"], r["效期"]))
    return result


def _count_statuses(rows: list[dict]) -> dict:
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["狀態"]] = counts.get(row["狀態"], 0) + 1
    return counts


# ------------------------------------------------------------------ 輸出

_DETAIL_COLUMNS = ["料號", "品名", "效期", "退貨授權數量", "入庫數量", "差異(授權-入庫)", "狀態"]


def _write_output(name1: str, name2: str, detail_rows: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = "核對摘要"
    summary["A1"] = "退貨物流核對摘要"
    summary["A1"].font = Font(bold=True, size=18)
    summary["A3"] = "產生時間"
    summary["B3"] = datetime.now()
    summary["B3"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary["A4"] = "來源檔一"
    summary["B4"] = name1
    summary["A5"] = "來源檔二"
    summary["B5"] = name2
    summary["A7"] = "核對總筆數"
    summary["B7"] = len(detail_rows)
    counts = _count_statuses(detail_rows)
    for offset, status in enumerate(["一致", "數量不符", "僅退貨授權", "僅入庫記錄"]):
        summary.cell(8 + offset, 1, status)
        summary.cell(8 + offset, 2, counts.get(status, 0))
    for ref in ("A3", "A4", "A5", "A7", "A8", "A9", "A10", "A11"):
        summary[ref].font = Font(bold=True)
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 48

    ws = wb.create_sheet("退貨核對明細")
    for col, header in enumerate(_DETAIL_COLUMNS, start=1):
        cell = ws.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    status_col = len(_DETAIL_COLUMNS)
    first_number_col = _DETAIL_COLUMNS.index("退貨授權數量") + 1
    for row_offset, record in enumerate(detail_rows, start=2):
        for col, header in enumerate(_DETAIL_COLUMNS, start=1):
            cell = ws.cell(row_offset, col, record[header])
            cell.border = border
            if col >= first_number_col and col != status_col:
                cell.number_format = "#,##0"
            if col == status_col:
                color = STATUS_COLORS.get(str(record[header]))
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
    ws.auto_filter.ref = (
        f"A1:{openpyxl.utils.get_column_letter(len(_DETAIL_COLUMNS))}{max(len(detail_rows) + 1, 1)}"
    )
    ws.freeze_panes = "A2"
    for col in range(1, len(_DETAIL_COLUMNS) + 1):
        letter = openpyxl.utils.get_column_letter(col)
        width = max(
            [len(str(_DETAIL_COLUMNS[col - 1])) * 2] +
            [min(len(str(r[_DETAIL_COLUMNS[col - 1]])), 40) for r in detail_rows[:200]] or [10]
        )
        ws.column_dimensions[letter].width = max(10, min(width + 2, 44))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
