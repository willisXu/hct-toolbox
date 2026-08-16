# -*- coding: utf-8 -*-
"""HCT 表格核對工具（由 VBA modHCTCompareTool 移植）。

比對「未出貨明細」與「銷貨單明細」兩份表格：
  - 數量核對：以 客戶+料號+效期 為鍵，加總兩邊數量比差異。兩份來源都沒有
    倉別欄位，所以這裡的鍵留在客戶層級（同料號同效期出給不同客戶的量不能
    合併）；效期取自批號欄，空白代表這列不在乎效期，仍會互相比對。
  - 訂單編號核對：以標準化訂單編號（取 - 前段）為鍵，比對兩邊出現情況。
兩份檔案順序不限，程式依欄位自動辨識類型。
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime

from .xlio import (
    build_header_map,
    first_sheet,
    is_excel_error_text,
    normalize_number,
    read_workbook,
    try_parse_date,
)

SOURCE_UNSHIPPED = "UNSHIPPED"
SOURCE_DELIVERY = "DELIVERY"
SOURCE_AMBIGUOUS = "AMBIGUOUS"

_UNSHIPPED_HEADERS = ["出貨客戶", "DR_料號", "交易序號/批號", "出貨數量", "客戶採購單編號"]
_DELIVERY_HEADERS = ["客戶簡稱", "品號", "批號", "銷貨數量", "網路訂單編號"]

STATUS_COLORS = {
    "一致": "E2EFDA",
    "兩邊皆有": "E2EFDA",
    "數量不符": "FFEB9C",
    "僅未出貨明細": "F4CCCC",
    "僅銷貨單": "BDD7EE",
}


class CompareError(ValueError):
    pass


@dataclass
class CompareResult:
    output_bytes: bytes
    output_name: str
    quantity_rows: int
    order_rows: int
    quantity_status_counts: dict
    order_status_counts: dict


def compare(
    data1: bytes, name1: str,
    data2: bytes, name2: str,
) -> CompareResult:
    rows1 = first_sheet(read_workbook(data1, name1))
    rows2 = first_sheet(read_workbook(data2, name2))
    headers1 = build_header_map(rows1[0]) if rows1 else {}
    headers2 = build_header_map(rows2[0]) if rows2 else {}

    type1 = _detect_type(headers1)
    type2 = _detect_type(headers2)
    _validate_types(type1, type2, name1, name2)

    if type1 == SOURCE_UNSHIPPED:
        unshipped, un_headers = rows1, headers1
        delivery, de_headers = rows2, headers2
    else:
        unshipped, un_headers = rows2, headers2
        delivery, de_headers = rows1, headers1

    quantity_rows = _build_quantity_compare(unshipped, un_headers, delivery, de_headers)
    order_rows = _build_order_compare(unshipped, un_headers, delivery, de_headers)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        output_bytes = _write_output(name1, name2, quantity_rows, order_rows)
    except Exception as exc:
        raise CompareError(f"建立輸出檔時發生錯誤：{exc}") from exc

    return CompareResult(
        output_bytes=output_bytes,
        output_name=f"HCT表格核對結果_{stamp}.xlsx",
        quantity_rows=len(quantity_rows),
        order_rows=len(order_rows),
        quantity_status_counts=_count_statuses(quantity_rows),
        order_status_counts=_count_statuses(order_rows),
    )


# ------------------------------------------------------------------ 正規化


def _normalize_text(value: object) -> str:
    """VBA 版 NormalizeText：IsError 視為空白、去尾端 .0（純數字時）。"""
    if value is None or is_excel_error_text(value):
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    text = str(value).strip()
    if len(text) > 2 and text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _normalize_customer(value: object) -> str:
    text = _normalize_text(value)
    if text[:4].upper() == "CUS-":
        space = text.find(" ")
        if space > 0:
            text = text[space + 1:].strip()
    return text


def _normalize_order_id(value: object) -> str:
    text = _normalize_text(value)
    dash = text.find("-")
    if dash > 0:
        text = text[:dash].strip()
    elif dash == 0:
        text = ""
    return text


def _normalize_expiry(value: object) -> tuple[str, str]:
    """批號欄即效期字串，回傳 (比對用 key, 顯示用文字)（同 return_compare）。

    兩份報表的同一個效期常寫成不同格式（20270501 / 2027-05-01 / 2027/5/1），
    直接拿原文當鍵會把同一批貨拆成兩列、雙雙變成「僅一邊有」的假差異。
    無法辨識為日期的批號保留原文（casefold 後加 raw: 前綴），不同批號不會
    被誤併；空白效期用空字串鍵，代表這列不在乎效期。
    """
    text = _normalize_text(value)
    if not text:
        return "", ""
    parsed = try_parse_date(value)
    if parsed:
        display = parsed.strftime("%Y-%m-%d")
        return display, display
    return f"raw:{text.casefold()}", text


def _cell(rows: list, row_index: int, headers: dict, name: str) -> object:
    index = headers.get(name.casefold())
    if index is None:
        return None
    row = rows[row_index]
    return row[index] if index < len(row) else None


def _detect_type(headers: dict) -> str:
    has_unshipped = all(h.casefold() in headers for h in _UNSHIPPED_HEADERS)
    has_delivery = all(h.casefold() in headers for h in _DELIVERY_HEADERS)
    if has_unshipped and has_delivery:
        return SOURCE_AMBIGUOUS
    if has_unshipped:
        return SOURCE_UNSHIPPED
    if has_delivery:
        return SOURCE_DELIVERY
    return ""


def _validate_types(type1: str, type2: str, name1: str, name2: str) -> None:
    if SOURCE_AMBIGUOUS in (type1, type2):
        raise CompareError("有檔案同時符合兩種表格欄位，無法自動判斷。")
    if not type1 or not type2:
        raise CompareError(
            "無法辨識其中一個檔案的欄位。\n"
            f"檔案一:{name1},類型:{type1 or '無法辨識'}\n"
            f"檔案二:{name2},類型:{type2 or '無法辨識'}\n\n"
            "未出貨明細需要欄位:" + "、".join(_UNSHIPPED_HEADERS) + "\n"
            "銷貨單明細需要欄位:" + "、".join(_DELIVERY_HEADERS)
        )
    if type1 == type2:
        raise CompareError("請選擇一份未出貨明細與一份銷貨單明細，不要選到同類型檔案。")


# ------------------------------------------------------------------ 核對


def _build_quantity_compare(unshipped, un_headers, delivery, de_headers) -> list[dict]:
    records: dict[tuple, dict] = {}

    def accumulate(rows, headers, customer_h, item_h, batch_h, qty_h, field):
        for row_index in range(1, len(rows)):
            customer = _normalize_customer(_cell(rows, row_index, headers, customer_h))
            item = _normalize_text(_cell(rows, row_index, headers, item_h))
            expiry_key, expiry_display = _normalize_expiry(
                _cell(rows, row_index, headers, batch_h))
            quantity = normalize_number(_cell(rows, row_index, headers, qty_h))
            if not (customer or item or expiry_display):
                continue
            key = (customer.casefold(), item.casefold(), expiry_key)
            record = records.setdefault(key, {
                "客戶": customer, "料號": item, "效期": expiry_display,
                "未出貨數量": 0.0, "銷貨數量": 0.0,
            })
            record[field] += quantity

    accumulate(unshipped, un_headers, "出貨客戶", "DR_料號", "交易序號/批號", "出貨數量", "未出貨數量")
    accumulate(delivery, de_headers, "客戶簡稱", "品號", "批號", "銷貨數量", "銷貨數量")

    result = []
    for record in records.values():
        un_qty, de_qty = record["未出貨數量"], record["銷貨數量"]
        record["差異(未出貨-銷貨)"] = un_qty - de_qty
        if un_qty == 0 and de_qty != 0:
            record["狀態"] = "僅銷貨單"
        elif un_qty != 0 and de_qty == 0:
            record["狀態"] = "僅未出貨明細"
        elif abs(un_qty - de_qty) < 1e-6:
            record["狀態"] = "一致"
        else:
            record["狀態"] = "數量不符"
        result.append(record)
    return result


def _build_order_compare(unshipped, un_headers, delivery, de_headers) -> list[dict]:
    records: dict[str, dict] = {}

    def accumulate(rows, headers, order_h, customer_h, qty_h, prefix, qty_field):
        for row_index in range(1, len(rows)):
            original = _normalize_text(_cell(rows, row_index, headers, order_h))
            order_id = _normalize_order_id(original)
            if not order_id:
                continue
            record = records.setdefault(order_id.casefold(), {
                "標準化訂單編號": order_id,
                "未出貨原始編號": [], "銷貨單原始編號": [],
                "未出貨客戶": [], "銷貨單客戶": [],
                "未出貨筆數": 0, "銷貨單筆數": 0,
                "未出貨數量": 0.0, "銷貨數量": 0.0,
            })
            customer = _normalize_customer(_cell(rows, row_index, headers, customer_h))
            quantity = normalize_number(_cell(rows, row_index, headers, qty_h))
            originals = record[f"{prefix}原始編號"]
            if original and original not in originals:
                originals.append(original)
            customers = record[f"{prefix}客戶"]
            if customer and customer not in customers:
                customers.append(customer)
            record[f"{prefix}筆數"] += 1
            record[qty_field] += quantity

    accumulate(unshipped, un_headers, "客戶採購單編號", "出貨客戶", "出貨數量", "未出貨", "未出貨數量")
    accumulate(delivery, de_headers, "網路訂單編號", "客戶簡稱", "銷貨數量", "銷貨單", "銷貨數量")

    result = []
    for record in records.values():
        record["未出貨原始編號"] = " / ".join(record["未出貨原始編號"])
        record["銷貨單原始編號"] = " / ".join(record["銷貨單原始編號"])
        record["未出貨客戶"] = " / ".join(record["未出貨客戶"])
        record["銷貨單客戶"] = " / ".join(record["銷貨單客戶"])
        record["差異(未出貨-銷貨)"] = record["未出貨數量"] - record["銷貨數量"]
        if record["未出貨筆數"] > 0 and record["銷貨單筆數"] > 0:
            record["狀態"] = "兩邊皆有"
        elif record["未出貨筆數"] > 0:
            record["狀態"] = "僅未出貨明細"
        else:
            record["狀態"] = "僅銷貨單"
        result.append(record)
    return result


def _count_statuses(rows: list[dict]) -> dict:
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["狀態"]] = counts.get(row["狀態"], 0) + 1
    return counts


# ------------------------------------------------------------------ 輸出

_QUANTITY_COLUMNS = ["客戶", "料號", "效期", "未出貨數量", "銷貨數量", "差異(未出貨-銷貨)", "狀態"]
_ORDER_COLUMNS = [
    "標準化訂單編號", "未出貨原始編號", "銷貨單原始編號", "未出貨客戶", "銷貨單客戶",
    "未出貨筆數", "銷貨單筆數", "未出貨數量", "銷貨數量", "差異(未出貨-銷貨)", "狀態",
]


def _write_output(name1: str, name2: str, quantity_rows: list, order_rows: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = "核對摘要"
    summary["A1"] = "HCT 表格核對摘要"
    summary["A1"].font = Font(bold=True, size=18)
    summary["A3"] = "產生時間"
    summary["B3"] = datetime.now()
    summary["B3"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary["A4"] = "來源檔一"
    summary["B4"] = name1
    summary["A5"] = "來源檔二"
    summary["B5"] = name2
    summary["A7"] = "數量核對總筆數"
    summary["B7"] = len(quantity_rows)
    q_counts = _count_statuses(quantity_rows)
    for offset, status in enumerate(["一致", "數量不符", "僅未出貨明細", "僅銷貨單"]):
        summary.cell(8 + offset, 1, status)
        summary.cell(8 + offset, 2, q_counts.get(status, 0))
    summary["D7"] = "訂單編號核對總筆數"
    summary["E7"] = len(order_rows)
    o_counts = _count_statuses(order_rows)
    for offset, status in enumerate(["兩邊皆有", "僅未出貨明細", "僅銷貨單"]):
        summary.cell(8 + offset, 4, status)
        summary.cell(8 + offset, 5, o_counts.get(status, 0))
    for ref in ("A3", "A4", "A5", "A7", "A8", "A9", "A10", "A11", "D7", "D8", "D9", "D10"):
        summary[ref].font = Font(bold=True)
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 48
    summary.column_dimensions["D"].width = 22
    summary.column_dimensions["E"].width = 12

    def write_table(title: str, columns: list, rows: list, status_col: int, first_number_col: int):
        ws = wb.create_sheet(title)
        for col, header in enumerate(columns, start=1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for row_offset, record in enumerate(rows, start=2):
            for col, header in enumerate(columns, start=1):
                cell = ws.cell(row_offset, col, record[header])
                cell.border = border
                if col >= first_number_col and col != status_col:
                    cell.number_format = "#,##0"
                if col == status_col:
                    color = STATUS_COLORS.get(str(record[header]))
                    if color:
                        cell.fill = PatternFill("solid", fgColor=color)
        ws.auto_filter.ref = (
            f"A1:{openpyxl.utils.get_column_letter(len(columns))}{max(len(rows) + 1, 1)}"
        )
        ws.freeze_panes = "A2"
        for col in range(1, len(columns) + 1):
            letter = openpyxl.utils.get_column_letter(col)
            width = max(
                [len(str(columns[col - 1])) * 2] +
                [min(len(str(r[columns[col - 1]])), 40) for r in rows[:200]] or [10]
            )
            ws.column_dimensions[letter].width = max(10, min(width + 2, 44))

    write_table("數量核對", _QUANTITY_COLUMNS, quantity_rows, status_col=7, first_number_col=4)
    write_table("訂單編號核對", _ORDER_COLUMNS, order_rows, status_col=11, first_number_col=6)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
