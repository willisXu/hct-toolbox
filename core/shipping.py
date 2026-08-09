# -*- coding: utf-8 -*-
"""HCT 出貨單轉換引擎（訂單 360 格式 / 調撥單 162 格式共用）。

由 VBA modHCTConverter360 / modHCTConverter162 移植（第 4 點為移植後新增）：
  1. 依「預計到貨日 + 出貨客戶 + 門市/倉儲 + 地址」判斷可否跨單合併；
     同收件條件有 >=2 種訂單類型且含「一般銷售訂單」時合併為一張送貨單。
  2. 同送貨單內相同「料號 + 效期」的明細數量加總。
  3. 輸出 36 欄 HCT 銷貨報表（欄位常數取自 HCT範本）。
  4. 虛擬倉（AD 欄）：來源報表有倉別/地點欄時逐列帶出 G 開頭代碼；
     沒有該欄（或該列空白）時，料號 9 開頭（陳列/宣傳品，存 G90 倉）
     輸出 G90 並附警告，其餘沿用範本預設值（G10）。
  5. 備忘錄來源可選「明細行（備忘錄）」或「主要（備忘錄 (主要)）」；
     明細行備忘錄內容與該列品名相同時（2026-08 新版報表 NetSuite
     會回填品名）視為系統雜訊，不併入備註。
  6. 合併方式可選「收件條件（原規則，第 1 點）」或「客戶採購單編號」；
     後者採購單編號相同的訂單即合併為一張送貨單（不看訂單類型），
     編號空白的列不合併，合併群組內收件資訊不一致時警告並以第一筆為準。
"""
from __future__ import annotations

import io
import re
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from .xlio import (
    build_header_map,
    first_sheet,
    is_blank_row,
    normalize_text,
    read_workbook,
    try_parse_date,
)

GENERAL_ORDER_TYPE = "一般銷售訂單"
PLATFORM_NAME = "DR.WU 達爾膚生醫科技股份有限公司"
PLATFORM_URL = "https://www.drwu.com"
PLATFORM_SERVICE_PHONE = "0800-083-999"
CARRIER_NAME = "OMS"
OUTPUT_COLUMN_COUNT = 36

MODE_ORDER = "order"       # 銷售訂單未出貨明細 (360)
MODE_TRANSFER = "transfer"  # 調撥單未出貨明細 (162)

# 備忘錄（DR_溫馨提醒）來源欄選項：NetSuite 報表同時有明細行的「備忘錄」
# 與主要層級的「備忘錄 (主要)」時，由使用者決定要吃哪一欄。
MEMO_SOURCE_LINE = "line"   # 明細行「備忘錄」
MEMO_SOURCE_MAIN = "main"   # 主要「備忘錄 (主要)」

# 跨單合併方式選項。
MERGE_BY_RECIPIENT = "recipient"       # 原規則：到貨日+客戶+門市+地址，且類型>=2 含一般銷售訂單
MERGE_BY_PURCHASE_ORDER = "po"         # 客戶採購單編號相同即合併

_MEMO_ALIASES = {
    MEMO_SOURCE_LINE: ("備忘錄", "備忘錄 (主要)"),
    MEMO_SOURCE_MAIN: ("備忘錄 (主要)", "備忘錄"),
}

# 每個標準欄名可對應多個候選別名，依序取第一個存在的欄
#（2026-07 NetSuite 報表改版：項目名稱→顯示名稱、備忘錄→備忘錄 (主要)）。
_ARRIVAL_ALIASES = ("預計到貨日期", "預計到貨日", "DR_預計到貨日", "預計送達日期", "預計送達日")
_ARRIVAL_SLOT_ALIASES = ("預計到貨時段", "DR_預計到貨時間", "預計到貨時間")

_ALIASES = {
    MODE_ORDER: {
        "項目名稱": ("顯示名稱",),
        "序號/批號": ("交易序號/批號",),
        "DR_預計到貨日期": _ARRIVAL_ALIASES,
        "DR_預計到貨時段": _ARRIVAL_SLOT_ALIASES,
    },
    MODE_TRANSFER: {
        "項目名稱": ("顯示名稱",),
        "序號/批號": ("交易序號/批號",),
        "DR_預計到貨日期": _ARRIVAL_ALIASES,
        "DR_預計到貨時段": _ARRIVAL_SLOT_ALIASES,
        "出貨客戶": ("目標地點",),
        "門市/倉儲": ("倉儲",),
        "門市/倉儲聯繫人": ("倉儲聯繫人",),
        "門市/倉儲電話": ("倉儲電話",),
        "門市/倉儲地址": ("倉儲地址",),
    },
}

# 虛擬倉（輸出第 30 欄）逐列判斷用：來源報表的倉別/地點候選欄名（非必要欄位），
# 依序取第一個存在的欄。欄位值只要含 G+兩碼數字（如「G10」「G30 出貨倉」）即可辨識。
_WAREHOUSE_ALIASES = ("虛擬倉", "倉別", "地點", "出貨倉", "出貨地點")
_WAREHOUSE_CODE_RE = re.compile(r"G\d{2}")
# 料號 9 開頭為陳列/宣傳品（如 902126060002 POYA 陳列物），實際存放 G90 倉。
_DISPLAY_MATERIAL_PREFIX = "9"
DISPLAY_MATERIAL_WAREHOUSE = "G90"

_REQUIRED_HEADERS = {
    MODE_ORDER: [
        "內部 ID", "文件編號",
        "銷售訂單類型", "日期", "項目", "項目名稱", "DR_料號",
        "序號/批號", "出貨客戶",
        "出貨數量", "門市/倉儲", "門市/倉儲聯繫人",
        "門市/倉儲電話", "門市/倉儲地址",
        "DR_預計到貨日期", "DR_預計到貨時段",
    ],
    MODE_TRANSFER: [
        "內部 ID", "文件編號", "日期", "項目",
        "出貨數量", "出貨客戶", "門市/倉儲", "門市/倉儲聯繫人",
        "門市/倉儲電話", "門市/倉儲地址",
    ],
}


@dataclass
class Shipment:
    merged: bool
    arrival_value: object          # date / str / None
    customer: str
    location: str
    contact: str
    phone: str
    address: str
    items: dict = field(default_factory=dict)          # key -> item dict（保序）
    documents: list = field(default_factory=list)
    sales_orders: list = field(default_factory=list)
    general_sales_orders: list = field(default_factory=list)
    types: list = field(default_factory=list)
    times: list = field(default_factory=list)
    purchase_orders: list = field(default_factory=list)
    reminders: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


@dataclass
class ConvertResult:
    output_bytes: bytes
    output_name: str
    shipments: int
    output_items: int
    problem_count: int
    input_rows: int
    skipped_rows: int
    merged_groups: int
    warnings: list
    problem_rows: list


class ConvertError(ValueError):
    pass


def convert(
    data: bytes, filename: str, mode: str, template_path: Path,
    memo_source: str = MEMO_SOURCE_LINE,
    merge_by: str = MERGE_BY_RECIPIENT,
) -> ConvertResult:
    rows = first_sheet(read_workbook(data, filename))
    return convert_rows(rows, mode, template_path, memo_source=memo_source, merge_by=merge_by)


def convert_rows(
    rows: list[list[object]], mode: str, template_path: Path,
    memo_source: str = MEMO_SOURCE_LINE,
    merge_by: str = MERGE_BY_RECIPIENT,
) -> ConvertResult:
    """轉換已讀取的表格資料（header 列 + 資料列），供檔案上傳與 NetSuite 直接抓取共用。"""
    if not rows:
        raise ConvertError("來源工作表沒有可轉換的表格資料。")
    if merge_by not in (MERGE_BY_RECIPIENT, MERGE_BY_PURCHASE_ORDER):
        raise ConvertError(f"未知的合併方式選項：{merge_by}")

    headers = build_header_map(rows[0])
    alias_notes = _apply_aliases(headers, mode, memo_source)
    missing = [h for h in _REQUIRED_HEADERS[mode] if h.casefold() not in headers]
    if missing:
        raise ConvertError("來源檔缺少必要欄位：\n" + "\n".join(f"- {h}" for h in missing))

    state = _build_shipments(rows, headers, mode, merge_by)
    state["warnings"] = alias_notes + state["warnings"]
    total_items = sum(len(s.items) for s in state["shipments"].values())
    if total_items == 0:
        raise ConvertError("沒有可輸出的有效品項。")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    mode_label = "訂單" if mode == MODE_ORDER else "調撥單"
    output_name = f"HCT銷貨報表_{mode_label}_{stamp}.xlsx"
    try:
        output_bytes = _write_output(state, mode, template_path)
    except Exception as exc:
        raise ConvertError(f"建立輸出檔時發生錯誤：{exc}") from exc

    return ConvertResult(
        output_bytes=output_bytes,
        output_name=output_name,
        shipments=len(state["shipments"]),
        output_items=total_items,
        problem_count=len(state["problem_rows"]),
        input_rows=state["input_rows"],
        skipped_rows=state["skipped_rows"],
        merged_groups=state["merged_groups"],
        warnings=state["warnings"],
        problem_rows=state["problem_rows"],
    )


# ------------------------------------------------------------------ 讀表輔助


def _apply_aliases(
    headers: dict[str, int], mode: str, memo_source: str = MEMO_SOURCE_LINE,
) -> list[str]:
    """套用別名對照；回傳別名/關鍵字比對備援產生的提醒訊息（給呼叫端併入警告）。"""
    for canonical, aliases in _ALIASES[mode].items():
        key = canonical.casefold()
        if key in headers:
            continue
        for alias in aliases:
            alias_key = alias.casefold()
            if alias_key in headers:
                headers[key] = headers[alias_key]
                break
    return _apply_memo_source(headers, memo_source) + _match_arrival_by_keyword(headers)


def _apply_memo_source(headers: dict[str, int], memo_source: str) -> list[str]:
    """依使用者選擇綁定備忘錄來源欄（來源檔本身有 DR_溫馨提醒 欄時以該欄為準）。

    選擇的欄不存在但另一欄存在時，退回另一欄並回傳提醒訊息。
    """
    candidates = _MEMO_ALIASES.get(memo_source)
    if candidates is None:
        raise ConvertError(f"未知的備忘錄來源選項：{memo_source}")
    key = "DR_溫馨提醒".casefold()
    if key in headers:
        return []
    preferred = candidates[0]
    for alias in candidates:
        alias_key = alias.casefold()
        if alias_key in headers:
            headers[key] = headers[alias_key]
            if alias != preferred:
                return [
                    f"來源檔沒有「{preferred}」欄，備忘錄改用「{alias}」欄，"
                    "請確認備忘錄來源選擇是否正確。"
                ]
            return []
    return []


# 關鍵字比對備援要排除語意不同的欄位（實際到貨日、上次到貨日等），
# 避免誤綁後整批指定送達日都用錯欄位。
_ARRIVAL_EXCLUDED_WORDS = ("實際", "上次", "歷史")


def _match_arrival_by_keyword(headers: dict[str, int]) -> list[str]:
    """到貨日／時段欄名在 saved search 常被改動，別名沒命中時退回關鍵字比對。"""
    notes: list[str] = []
    for canonical, keyword, excluded in (
        ("DR_預計到貨日期", "到貨日", ("時段", "時間") + _ARRIVAL_EXCLUDED_WORDS),
        ("DR_預計到貨時段", "到貨時", _ARRIVAL_EXCLUDED_WORDS),
    ):
        key = canonical.casefold()
        if key in headers:
            continue
        for header, index in list(headers.items()):
            if keyword in header and not any(word in header for word in excluded):
                headers[key] = index
                notes.append(
                    f"找不到「{canonical}」欄，改用關鍵字比對到的來源欄「{header}」，"
                    "請確認來源報表欄位是否正確。"
                )
                break
    return notes


def _cell(row: list, headers: dict[str, int], name: str) -> object:
    index = headers.get(name.casefold())
    if index is None or index >= len(row):
        return None
    return row[index]


def _add_distinct(values: list, text: str) -> None:
    if text and text not in values:
        values.append(text)


def _normalize_date_key(value: object) -> str:
    parsed = try_parse_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed else ""


def _normalize_arrival(value: object, merge_hint: bool = False) -> tuple[object, str]:
    """merge_hint 只有訂單模式（真的有跨單合併邏輯）才傳 True；
    M00 與調撥單模式不合併，空白警告不該暗示「填了就會合併」。"""
    raw = normalize_text(value)
    if not raw:
        if merge_hint:
            return None, "預計到貨日期空白；此 SO 不跨單合併。"
        return None, "預計到貨日期空白。"
    parsed = try_parse_date(value)
    if parsed:
        return parsed, ""
    return raw, f"預計到貨日期無法辨識：{raw}"


def _normalize_expiry(batch_text: str) -> tuple[str, object, str]:
    """回傳 (expiry_key, expiry_value, warning)。批號本身即效期字串。

    解析失敗時 key 保留原始批號文字（加 raw: 前綴與日期 key 區隔），
    避免同料號下兩個都無法解析的「不同」批號被靜默合併成同一筆。
    """
    if not batch_text:
        return "", None, ""
    parsed = try_parse_date(batch_text)
    if parsed:
        return parsed.strftime("%Y-%m-%d"), parsed, ""
    return f"raw:{batch_text.casefold()}", None, f"效期無法辨識，已留白：{batch_text}"


# 「項目」欄的「料號直接連品名」樣態（2026-07 備品列新增）：開頭一串數字
# 料號後面直接接品名文字，如 902224100008ㄈ型試用品壓克力架(圓管)。
_ITEM_CODE_NAME_RE = re.compile(r"^(\d{6,})\s*(\D.*)$")


def _material_and_name(dr_material: object, item_text: object) -> tuple[str, str]:
    """從 DR_料號／項目 解析 (料號, 品名備援)。

    DR_料號空白時退回「項目」欄，該欄有三種樣態：
    料號_序號（取底線前段）、純料號、料號直接連品名（拆開後品名當備援）。
    """
    material = normalize_text(dr_material)
    if material:
        return material, ""
    fallback = normalize_text(item_text)
    matched = _ITEM_CODE_NAME_RE.match(fallback)
    if "_" in fallback:
        head = fallback.split("_", 1)[0]
        if head:
            fallback = head
    elif matched:
        return matched.group(1), matched.group(2).strip()
    return fallback.strip(), ""


_REMINDER_NOISE_NOTE = (
    "部分明細列的備忘錄內容與品名相同（NetSuite 系統帶入的預設值），未併入備註；"
    "如需訂單備註請將備忘錄來源改選「主要備忘錄」。"
)


def _reminder_text(row: list, headers: dict[str, int]) -> tuple[str, bool]:
    """讀該列備忘錄（DR_溫馨提醒）欄，回傳 (內容, 是否為雜訊)。

    2026-08 新版報表的明細行「備忘錄」被 NetSuite 回填成品名，
    與該列項目名稱相同時視為雜訊忽略，避免備註被一長串品名塞爆。
    """
    text = normalize_text(_cell(row, headers, "DR_溫馨提醒"))
    if not text:
        return "", False
    name = normalize_text(_cell(row, headers, "項目名稱"))
    if name and text == name:
        return "", True
    return text, False


BATCH_STATUS_OK = "批號賦予成功"


def _batch_status_warning(row: list, headers: dict[str, int], row_number: int) -> str:
    """調撥單報表的 DR_程式執行狀態非「批號賦予成功」時給警告（欄位不存在則略過）。"""
    if "DR_程式執行狀態".casefold() not in headers:
        return ""
    status = normalize_text(_cell(row, headers, "DR_程式執行狀態"))
    if status and status != BATCH_STATUS_OK:
        return f"第 {row_number} 列批號狀態為「{status}」，批號/效期可能尚未確認。"
    return ""


def _resolve_warehouse(row: list, headers: dict[str, int], material: str) -> tuple[str, str]:
    """判斷該列的虛擬倉，回傳 (倉別代碼, 警告)。

    代碼空字串表示交由範本預設值（G10）決定。來源有倉別/地點欄時以欄位值
    為準；沒有該欄或該列空白時退回料號規則（9 開頭 → G90）。
    """
    for name in _WAREHOUSE_ALIASES:
        if name.casefold() not in headers:
            continue
        raw = normalize_text(_cell(row, headers, name))
        if raw:
            matched = _WAREHOUSE_CODE_RE.search(raw.upper())
            if matched:
                return matched.group(0), ""
            return "", f"倉別「{raw}」無法辨識出 G 開頭倉別代碼，虛擬倉改用範本預設值。"
        break
    if material.startswith(_DISPLAY_MATERIAL_PREFIX):
        return (
            DISPLAY_MATERIAL_WAREHOUSE,
            f"料號 {material} 為 9 開頭（陳列/宣傳品），"
            f"虛擬倉輸出 {DISPLAY_MATERIAL_WAREHOUSE}，請確認。",
        )
    return "", ""


def _positive_quantity(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        quantity = float(str(value).replace(",", "").strip())
    except ValueError:
        return None
    return quantity if quantity > 0 else None


def _phone_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    return normalize_text(value)


def _split_phone(raw_phone: str) -> tuple[str, str]:
    compact = raw_phone.strip()
    for ch in ("-", " ", "(", ")", "\xa0"):
        compact = compact.replace(ch, "")
    digits = compact.lstrip("+")
    if digits.startswith("886"):
        digits = "0" + digits[3:]
    if digits and not digits.startswith("0"):
        digits = "0" + digits
    if not digits:
        return "", ""
    if digits.startswith("09"):
        return "", digits
    return digits, ""


def _split_postal(raw_address: str) -> tuple[str, str]:
    clean = raw_address.strip()
    if len(clean) >= 5 and clean[:5].isdigit():
        return clean[:5], clean[5:].strip()
    if len(clean) >= 3 and clean[:3].isdigit():
        return clean[:3], clean[3:].strip()
    return "", clean


# ------------------------------------------------------------------ 主轉換


def _build_shipments(
    rows: list, headers: dict[str, int], mode: str,
    merge_by: str = MERGE_BY_RECIPIENT,
) -> dict:
    warnings: list[str] = []
    problem_rows: list[dict] = []
    input_rows = sum(1 for row in rows[1:] if not is_blank_row(row))
    skipped_rows = 0

    def candidate_key_for(row: list) -> tuple:
        return (
            _normalize_date_key(_cell(row, headers, "DR_預計到貨日期")),
            normalize_text(_cell(row, headers, "出貨客戶")),
            normalize_text(_cell(row, headers, "門市/倉儲")),
            normalize_text(_cell(row, headers, "門市/倉儲地址")),
        )

    def document_for(row: list) -> str:
        document = normalize_text(_cell(row, headers, "文件編號"))
        return document or "ID-" + normalize_text(_cell(row, headers, "內部 ID"))

    # 合併判斷的前置掃描：
    #   收件條件模式 → 同鍵的訂單類型集合（>=2 種且含一般銷售訂單才合併）；
    #   採購單模式   → 同採購單編號的文件編號集合（>=2 張單才算合併群組）。
    candidate_types: dict[tuple, list[str]] = {}
    candidate_has_general: set = set()
    po_documents: dict[str, set] = {}

    if merge_by == MERGE_BY_PURCHASE_ORDER:
        if "客戶採購單編號".casefold() not in headers:
            _add_distinct(
                warnings,
                "來源檔沒有「客戶採購單編號」欄，無法以採購單合併，全部各自成單。",
            )
        for row in rows[1:]:
            if is_blank_row(row):
                continue
            po = normalize_text(_cell(row, headers, "客戶採購單編號"))
            if po:
                po_documents.setdefault(po, set()).add(document_for(row))
        merged_groups = sum(1 for docs in po_documents.values() if len(docs) >= 2)
    else:
        for row in rows[1:]:
            key = candidate_key_for(row)
            if not all(key):
                continue
            order_type = normalize_text(_cell(row, headers, "銷售訂單類型"))
            types = candidate_types.setdefault(key, [])
            _add_distinct(types, order_type)
            if order_type.casefold() == GENERAL_ORDER_TYPE.casefold():
                candidate_has_general.add(key)

        merged_groups = sum(
            1
            for key, types in candidate_types.items()
            if len(types) >= 2 and key in candidate_has_general
        )

    shipments: dict[tuple, Shipment] = {}

    for row_number, row in enumerate(rows[1:], start=2):
        if is_blank_row(row):
            continue

        document = normalize_text(_cell(row, headers, "文件編號"))
        if not document:
            document = "ID-" + normalize_text(_cell(row, headers, "內部 ID"))
            _add_distinct(warnings, f"第 {row_number} 列文件編號空白，改用 {document}")
        sales_order = normalize_text(_cell(row, headers, "銷售訂單單號"))
        if not sales_order:
            sales_order = document
            if "銷售訂單單號".casefold() in headers:
                _add_distinct(warnings, f"第 {row_number} 列銷售訂單單號空白，改用 {sales_order}")

        if merge_by == MERGE_BY_PURCHASE_ORDER:
            po = normalize_text(_cell(row, headers, "客戶採購單編號"))
            mergeable = bool(po)
            is_merged = mergeable and len(po_documents.get(po, ())) >= 2
            shipment_key = ("P", po) if is_merged else ("S", sales_order, po)
            merge_blank_warning = "客戶採購單編號空白；此 SO 不跨單合併。"
        else:
            key = candidate_key_for(row)
            mergeable = all(key)
            is_merged = (
                mergeable
                and key in candidate_types
                and len(candidate_types[key]) >= 2
                and key in candidate_has_general
            )
            shipment_key = ("M", *key) if is_merged else ("S", sales_order, *key)
            merge_blank_warning = "合併鍵不完整；此 SO 不跨單合併。"

        shipment = shipments.get(shipment_key)
        if shipment is None:
            arrival_value, arrival_warning = _normalize_arrival(
                _cell(row, headers, "DR_預計到貨日期"),
                merge_hint=mode == MODE_ORDER,
            )
            shipment = Shipment(
                merged=is_merged,
                arrival_value=arrival_value,
                customer=normalize_text(_cell(row, headers, "出貨客戶")),
                location=normalize_text(_cell(row, headers, "門市/倉儲")),
                contact=normalize_text(_cell(row, headers, "門市/倉儲聯繫人")),
                phone=_phone_to_text(_cell(row, headers, "門市/倉儲電話")),
                address=normalize_text(_cell(row, headers, "門市/倉儲地址")),
            )
            if arrival_warning:
                _add_distinct(shipment.warnings, arrival_warning)
                _add_distinct(warnings, arrival_warning)
            if mode == MODE_ORDER and not mergeable:
                _add_distinct(shipment.warnings, merge_blank_warning)
                _add_distinct(warnings, merge_blank_warning)
            shipments[shipment_key] = shipment
        elif merge_by == MERGE_BY_PURCHASE_ORDER and shipment.merged:
            # 收件條件模式的合併鍵本身就含收件資訊，同鍵必一致；
            # 採購單模式合併鍵只有採購單編號，收件資訊不一致要提醒（輸出以第一筆為準）。
            mismatches = [
                label
                for label, first, current in (
                    ("出貨客戶", shipment.customer, normalize_text(_cell(row, headers, "出貨客戶"))),
                    ("門市/倉儲地址", shipment.address, normalize_text(_cell(row, headers, "門市/倉儲地址"))),
                    (
                        "預計到貨日期",
                        _normalize_date_key(shipment.arrival_value),
                        _normalize_date_key(_cell(row, headers, "DR_預計到貨日期")),
                    ),
                )
                if current and current != first
            ]
            if mismatches:
                # 訊息不帶列號：同一張採購單的同一種不一致只提醒一次，
                # 避免每列各發一則把警告區塞爆。
                text = (
                    f"採購單 {shipment_key[1]} 合併的訂單「{'、'.join(mismatches)}」不一致，"
                    "輸出以第一筆為準，請確認是否應該合併。"
                )
                _add_distinct(shipment.warnings, text)
                _add_distinct(warnings, text)

        skipped_rows += _add_row_to_shipment(
            shipment, row, row_number, headers, document, sales_order,
            warnings, problem_rows, mode,
        )

    return {
        "shipments": shipments,
        "warnings": warnings,
        "problem_rows": problem_rows,
        "input_rows": input_rows,
        "skipped_rows": skipped_rows,
        "merged_groups": merged_groups,
    }


def _add_row_to_shipment(
    shipment: Shipment,
    row: list,
    row_number: int,
    headers: dict[str, int],
    document: str,
    sales_order: str,
    warnings: list,
    problem_rows: list,
    mode: str,
) -> int:
    """把一列加進送貨單；回傳跳過列數 (0 或 1)。"""
    _add_distinct(shipment.documents, document)
    _add_distinct(shipment.sales_orders, sales_order)

    order_type = normalize_text(_cell(row, headers, "銷售訂單類型"))
    _add_distinct(shipment.types, order_type)
    if order_type.casefold() == GENERAL_ORDER_TYPE.casefold():
        _add_distinct(shipment.general_sales_orders, sales_order)

    _add_distinct(shipment.times, normalize_text(_cell(row, headers, "DR_預計到貨時段")))
    _add_distinct(shipment.purchase_orders, normalize_text(_cell(row, headers, "客戶採購單編號")))
    reminder, reminder_is_noise = _reminder_text(row, headers)
    _add_distinct(shipment.reminders, reminder)
    if reminder_is_noise:
        _add_distinct(warnings, _REMINDER_NOISE_NOTE)

    batch_value = normalize_text(_cell(row, headers, "序號/批號"))
    item_text = normalize_text(_cell(row, headers, "項目"))
    material, name_fallback = _material_and_name(_cell(row, headers, "DR_料號"), item_text)
    product_name = normalize_text(_cell(row, headers, "項目名稱")) or name_fallback

    status_warning = _batch_status_warning(row, headers, row_number)
    if status_warning:
        _add_distinct(warnings, status_warning)
        _add_distinct(shipment.warnings, status_warning)

    def record_problem(reason: str) -> None:
        problem_rows.append({
            "來源列": row_number,
            "文件編號": document,
            "銷售訂單單號": sales_order,
            "銷售訂單類型": order_type,
            "項目": item_text,
            "DR_料號": normalize_text(_cell(row, headers, "DR_料號")),
            "項目名稱": product_name,
            "序號/批號": batch_value,
            "出貨數量": normalize_text(_cell(row, headers, "出貨數量")),
            "出貨客戶": shipment.customer,
            "門市/倉儲": shipment.location,
            "門市/倉儲地址": shipment.address,
            "DR_預計到貨日期": normalize_text(_cell(row, headers, "DR_預計到貨日期")),
            "排除原因": reason,
        })

    if not material:
        text = f"第 {row_number} 列無有效料號，已跳過。"
        _add_distinct(warnings, text)
        _add_distinct(shipment.warnings, text)
        record_problem("無有效料號")
        return 1

    quantity = _positive_quantity(_cell(row, headers, "出貨數量"))
    if quantity is None:
        text = f"第 {row_number} 列數量無效，已跳過。"
        _add_distinct(warnings, text)
        _add_distinct(shipment.warnings, text)
        record_problem("數量無效")
        return 1

    expiry_key, expiry_value, expiry_warning = _normalize_expiry(batch_value)
    if expiry_warning:
        text = f"第 {row_number} 列：{expiry_warning}"
        _add_distinct(warnings, text)
        _add_distinct(shipment.warnings, text)

    warehouse, warehouse_warning = _resolve_warehouse(row, headers, material)
    if warehouse_warning:
        _add_distinct(warnings, warehouse_warning)
        _add_distinct(shipment.warnings, warehouse_warning)

    item_key = (material, expiry_key)
    item = shipment.items.get(item_key)
    if item is not None:
        item["quantity"] += quantity
        if item["warehouse"] != warehouse:
            text = (
                f"第 {row_number} 列料號 {material} 倉別（{warehouse or '預設'}）"
                f"與同單前列（{item['warehouse'] or '預設'}）不一致，輸出以前列為準。"
            )
            _add_distinct(warnings, text)
            _add_distinct(shipment.warnings, text)
    else:
        shipment.items[item_key] = {
            "material": material,
            "product_name": product_name,
            "batch": batch_value,
            "expiry_value": expiry_value,
            "expiry_key": expiry_key,
            "quantity": quantity,
            "warehouse": warehouse,
        }
    return 0


# ------------------------------------------------------------------ 輸出


def _build_note(shipment: Shipment) -> str:
    return "|".join((
        "、".join(shipment.times),
        "、".join(shipment.purchase_orders),
        "、".join(shipment.reminders),
    ))


def _delivery_number(shipment: Shipment) -> str:
    return "、".join(shipment.sales_orders)


def _order_number(shipment: Shipment, mode: str) -> str:
    if shipment.merged:
        # 採購單合併的群組可能完全沒有一般銷售訂單（如全是備品出貨），
        # 此時退回列出所有 SO，不能輸出空白訂單編號。
        if not shipment.general_sales_orders:
            return "、".join(shipment.sales_orders)
        if mode == MODE_ORDER:
            return shipment.general_sales_orders[0]
        return "、".join(shipment.general_sales_orders)
    return "、".join(shipment.sales_orders)


_TEXT_COLUMNS = (3, 5, 7, 11, 13, 15)
# 對齊 HCT 可匯入範本（260716銷貨-1 NS-水嫩.xlsx）：
#   指定送達日(29) = 文字字串 YYYY-MM-DD；效期(33) = 日期值 + Excel 內建
#   短日期格式（格式代碼 14，openpyxl 表示為 mm-dd-yy，zh-TW 顯示 yyyy/m/d）。
_EXPIRY_COLUMN = 33
_EXPIRY_FORMAT = "mm-dd-yy"
# 第 30 欄（虛擬倉）的範本值只在 _resolve_warehouse 判斷不出倉別時當預設值用。
_TEMPLATE_VALUE_COLUMNS = (4, 20, 21, 22, 23, 25, 26, 27, 30, 31, 32, 34)


def _arrival_text(value: object) -> str:
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return "" if value is None else str(value)


def _write_output(state: dict, mode: str, template_path: Path) -> bytes:
    import openpyxl

    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb.active

    template_values = {
        col: template_ws.cell(2, col).value for col in _TEMPLATE_VALUE_COLUMNS
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工作表1"

    for col in range(1, OUTPUT_COLUMN_COUNT + 1):
        header_cell = template_ws.cell(1, col)
        cell = ws.cell(1, col, header_cell.value)
        cell.font = copy(header_cell.font)
        cell.fill = copy(header_cell.fill)
        cell.border = copy(header_cell.border)
        cell.alignment = copy(header_cell.alignment)
        letter = openpyxl.utils.get_column_letter(col)
        width = template_ws.column_dimensions[letter].width
        if width:
            ws.column_dimensions[letter].width = width

    row_index = 1
    for shipment in state["shipments"].values():
        postal, address_text = _split_postal(shipment.address)
        daytime, mobile = _split_phone(shipment.phone)
        note = _build_note(shipment)
        for item in shipment.items.values():
            row_index += 1
            values: list[object] = [None] * (OUTPUT_COLUMN_COUNT + 1)
            values[1] = _delivery_number(shipment)
            values[2] = _order_number(shipment, mode)
            values[3] = item["material"]
            values[4] = template_values[4]
            values[5] = item["material"]
            values[6] = item["product_name"]
            values[7] = item["batch"]
            values[8] = item["quantity"]
            values[9] = ""
            values[10] = shipment.contact
            values[11] = postal
            values[12] = address_text
            values[13] = daytime
            values[14] = ""
            values[15] = mobile
            values[16] = PLATFORM_NAME
            values[17] = PLATFORM_URL
            values[18] = PLATFORM_SERVICE_PHONE
            values[19] = note
            for col in (20, 21, 22, 23, 25, 26, 27, 31, 32, 34):
                values[col] = template_values[col]
            values[24] = CARRIER_NAME
            values[28] = ""
            values[29] = _arrival_text(shipment.arrival_value)
            values[30] = item["warehouse"] or template_values[30]
            values[33] = item["expiry_value"]
            values[35] = shipment.customer
            values[36] = 1
            for col in range(1, OUTPUT_COLUMN_COUNT + 1):
                cell = ws.cell(row_index, col)
                value = values[col]
                if isinstance(value, date):
                    cell.value = value
                    cell.number_format = _EXPIRY_FORMAT if col == _EXPIRY_COLUMN else "yyyy-mm-dd"
                elif col in _TEXT_COLUMNS:
                    cell.number_format = "@"
                    cell.value = "" if value is None else str(value)
                else:
                    cell.value = value

    ws.auto_filter.ref = f"A1:AJ{row_index}"
    ws.freeze_panes = "A2"

    _write_problem_sheet(wb, state["problem_rows"])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


_PROBLEM_HEADERS = [
    "來源列", "文件編號", "銷售訂單單號", "銷售訂單類型",
    "項目", "DR_料號", "項目名稱", "序號/批號", "出貨數量", "出貨客戶",
    "門市/倉儲", "門市/倉儲地址", "DR_預計到貨日期", "排除原因",
]


def _write_problem_sheet(wb, problem_rows: list) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet("有問題訂單")
    for col, header in enumerate(_PROBLEM_HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
    ws.column_dimensions["H"].width = 16
    for col_cells in ws.iter_cols(min_col=8, max_col=8, min_row=2):
        for cell in col_cells:
            cell.number_format = "@"
    if problem_rows:
        for row_offset, record in enumerate(problem_rows, start=2):
            for col, header in enumerate(_PROBLEM_HEADERS, start=1):
                cell = ws.cell(row_offset, col, record.get(header, ""))
                if header == "序號/批號":
                    cell.number_format = "@"
    else:
        ws.cell(2, 1, "（本次沒有無法轉換的明細）")
