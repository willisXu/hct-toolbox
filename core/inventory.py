# -*- coding: utf-8 -*-
"""HCT／代工廠／M00 × NetSuite 庫存核對(由 VBA modInventory* 模組移植)。

  - NetSuite 報表欄位：地點、到期日、DR_料號、項目計數 總和、數量 總和
  - NetSuite 批號版報表欄位：料號、品名、倉別代碼、倉別名稱、批號、在庫數量
  - HCT 報表欄位：儲區類別、有效日期、客戶產品編號、可出數量、庫存數量
  - 代工廠報表欄位：庫存日期、倉別、品號、品名、批號、庫存數量
  - M00 報表(電商物流 InventorySummaryReport)：取「庫存詳情」分頁，
    欄位 SKU、商品名稱、有效日期、數量、組合保留、批號；報表沒有倉別欄，
    一律視為 M00 倉，庫存量＝數量＋組合保留(單品數量庫存，已用
    「庫存總表-單品」的庫存總數驗證)，同時當作可用量與總庫存量。
  - HCT 對帳只核對 G00/G10/G30/G40/G80/G90 倉；代工廠對帳只核對 D 開頭
    代工廠倉(D01 凱芬妮、D03 詠麗…)，「合計／總計」小計列列入排除數；
    M00 對帳只核對 M00 開頭倉別(各 M00 開頭代碼收斂成 M00 核對鍵)。
    非核對範圍的倉別一律列入排除數。
  - 代工廠報表只有一個「庫存數量」欄，同時當作可用量與總庫存量(同 293 格式)。
  - 以「倉別+料號+到期日」做日期明細核對，另以「倉別+料號」做料號彙總核對。
  - 差異基準：HCT／代工廠／M00－NetSuite。
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime

from .xlio import is_excel_error_text, read_workbook, try_parse_date

SYSTEM_HCT = "HCT"
SYSTEM_CONTRACT = "代工廠"
SYSTEM_M00 = "M00"
SYSTEM_NETSUITE = "NetSuite"
APPROVED_LOCATIONS = {"G00", "G10", "G30", "G40", "G80", "G90"}
CONTRACT_LOCATION_PREFIX = "D"
M00_LOCATION_PREFIX = "M00"

STATUS_MATCH = "完全一致"
STATUS_AVAILABLE_DIFF = "僅可用量差異"
STATUS_TOTAL_DIFF = "僅總庫存差異"
STATUS_BOTH_DIFF = "兩者皆不同"
STATUS_HCT_ONLY = "僅 HCT 存在"
STATUS_NS_ONLY = "僅 NetSuite 存在"


def _ext_only_status(ext_label: str) -> str:
    return f"僅 {ext_label} 存在"


def status_order(ext_label: str = SYSTEM_HCT) -> list[str]:
    """依外部系統(HCT／代工廠／M00)產生狀態清單；「僅 X 存在」隨系統改名。"""
    return [
        STATUS_MATCH, STATUS_AVAILABLE_DIFF, STATUS_TOTAL_DIFF,
        STATUS_BOTH_DIFF, _ext_only_status(ext_label), STATUS_NS_ONLY,
    ]


ALL_STATUSES = status_order()

_NS_HEADERS = ["地點", "到期日", "DR_料號", "項目計數 總和", "數量 總和"]
# 業務助理版 NetSuite 庫存報表（293 格式）：料號取「項目」底線前段，
# 只有一個數量欄「庫存數 總和」，同時當作可用量與總庫存量。
_NS293_HEADERS = ["地點", "到期日", "項目", "庫存數 總和"]
# 物流核對版 NetSuite 庫存報表（663 格式）：料號用 DR_料號（空白時取「項目」
# 底線前段），到期日在「庫存編號」欄，可用量=可用、總庫存量=在庫量。
_NS663_HEADERS = ["項目", "DR_料號", "地點", "庫存編號", "在庫量", "可用"]
# 批號版 NetSuite 庫存報表：倉別直接是代碼欄(G10、G80…，另有「倉別名稱」僅供
# 參考不參與核對)，效期放「批號」欄(yyyymmdd)，只有一個「在庫數量」欄，同時
# 當作可用量與總庫存量(同 293／代工廠)。核對鍵一樣是 倉別＋料號＋效期。
# 必要欄位刻意不含「倉別名稱」：代工廠報表也有 倉別名稱／品名／批號，
# 靠 料號＋倉別代碼＋在庫數量 三欄才能跟它區隔開，避免同時命中兩種格式。
_NS_LOT_HEADERS = ["料號", "品名", "倉別代碼", "批號", "在庫數量"]
_HCT_HEADERS = ["客戶產品編號", "有效日期", "儲區類別", "可出數量", "庫存數量"]
# 代工廠庫存核對報表：只有一個「庫存數量」欄，同時當作可用量與總庫存量；
# 「批號」放到期日欄位(目前多為空白＝無到期日)，「合計／總計」小計列會被排除。
_CONTRACT_HEADERS = ["庫存日期", "倉別", "品號", "品名", "庫存數量"]
# M00 電商物流 InventorySummaryReport 的「庫存詳情」分頁：沒有倉別欄
# (一律視為 M00 倉)，庫存量＝數量＋組合保留。這組欄位只有庫存詳情分頁
# 同時具備，不會誤中同檔案的庫存總表／待入庫清單等其他分頁。
_M00_HEADERS = ["SKU", "商品名稱", "有效日期", "數量", "組合保留", "批號"]
# M00 報表「有效日期」的無到期日符號（贈品/週邊等）：涵蓋半形/全形各種
# 破折號寫法，只認 ASCII 的 "-" 會讓其他寫法的列被誤判成日期異常整列排除。
_M00_NO_EXPIRY_MARKS = {"-", "—", "–", "‑", "－"}

SYSTEM_NETSUITE_293 = "NetSuite293"
SYSTEM_NETSUITE_663 = "NetSuite663"
SYSTEM_NETSUITE_LOT = "NetSuite批號版"


class InventoryError(ValueError):
    pass


@dataclass
class SourceStats:
    file_name: str
    sheet_name: str = ""
    rows_read: int = 0
    valid_rows: int = 0
    excluded_rows: int = 0
    anomaly_rows: int = 0


@dataclass
class InventoryResult:
    output_bytes: bytes
    output_name: str
    detail_rows: list
    item_rows: list
    anomalies: list
    hct_stats: SourceStats
    ns_stats: SourceStats
    detail_status_counts: dict = field(default_factory=dict)
    item_status_counts: dict = field(default_factory=dict)
    ext_label: str = SYSTEM_HCT
    statuses: list = field(default_factory=lambda: list(ALL_STATUSES))


# ------------------------------------------------------------------ 正規化


def _normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    return text.replace("\xa0", " ").replace("　", " ").strip()


def _normalize_item(value: object) -> str:
    if is_excel_error_text(value):
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return "" if value is None else str(value).strip()


def _normalize_location(value: object) -> str:
    if is_excel_error_text(value):
        return ""
    text = ("" if value is None else str(value)).strip().upper()
    if "_" in text:
        text = text.split("_", 1)[0]
    return text.strip()


def _normalize_date(value: object) -> tuple[str, bool]:
    """回傳 (yyyymmdd 或空字串, 是否有效)。空值視為有效的「無到期日」。

    日期解析交給 xlio.try_parse_date（datetime、Excel 序號、數字型別的
    YYYYMMDD、各種文字格式都涵蓋），避免自行維護一份分歧的解析邏輯。
    """
    if value is None or str(value).strip() == "":
        return "", True
    parsed = try_parse_date(value)
    if parsed:
        return parsed.strftime("%Y%m%d"), True
    return "", False


def _normalize_quantity(value: object) -> tuple[float, bool]:
    if value is None:
        return 0.0, True
    if isinstance(value, bool):
        return 0.0, False
    if isinstance(value, (int, float)):
        return float(value), True
    text = str(value).strip()
    if not text:
        return 0.0, True
    text = text.replace(",", "")
    try:
        return float(text), True
    except ValueError:
        return 0.0, False


def _expiry_output(expiry_key: str) -> object:
    if not expiry_key:
        return "無到期日"
    return date(int(expiry_key[:4]), int(expiry_key[4:6]), int(expiry_key[6:]))


# ------------------------------------------------------------------ 匯入


def _find_matching_sheet(book: dict, required: list[str]):
    """在所有工作表前 10 列內找欄位列；回傳 (sheet_name, rows, header_row_idx, header_map)。"""
    for sheet_name, rows in book.items():
        limit = min(len(rows), 10)
        for row_index in range(limit):
            header_map: dict[str, int] = {}
            for col_index, value in enumerate(rows[row_index]):
                text = _normalize_header(value)
                if text and text.casefold() not in header_map:
                    header_map[text.casefold()] = col_index
            if all(h.casefold() in header_map for h in required):
                return sheet_name, rows, row_index, header_map
    return None


def detect_source(data: bytes, filename: str) -> str:
    book = read_workbook(data, filename)
    matches = []
    if _find_matching_sheet(book, _NS_HEADERS):
        matches.append(SYSTEM_NETSUITE)
    if _find_matching_sheet(book, _NS663_HEADERS):
        matches.append(SYSTEM_NETSUITE_663)
    if _find_matching_sheet(book, _NS293_HEADERS):
        matches.append(SYSTEM_NETSUITE_293)
    if _find_matching_sheet(book, _NS_LOT_HEADERS):
        matches.append(SYSTEM_NETSUITE_LOT)
    if _find_matching_sheet(book, _HCT_HEADERS):
        matches.append(SYSTEM_HCT)
    if _find_matching_sheet(book, _CONTRACT_HEADERS):
        matches.append(SYSTEM_CONTRACT)
    if _find_matching_sheet(book, _M00_HEADERS):
        matches.append(SYSTEM_M00)
    if len(matches) > 1:
        raise InventoryError(
            f"{filename}:同時符合多種報表格式（{'、'.join(matches)}），無法自動判斷，"
            "請確認欄位名稱是否重複或誤植。"
        )
    return matches[0] if matches else ""


def _import_source(data: bytes, filename: str, system: str,
                   detail: dict, item_totals: dict, anomalies: list,
                   stats: SourceStats, location_ok, location_key=None) -> None:
    """location_key：通過 location_ok 的倉別轉成核對鍵（預設原樣）。
    M00 對帳用它把 NetSuite 側各種 M00 開頭代碼（M00倉、M001…）收斂成
    「M00」，跟 M00 報表側寫死的倉別對得上；否則前綴過濾放進來的列
    會因為鍵不同而變成假的「僅單邊存在」差異。"""
    book = read_workbook(data, filename)
    if system == SYSTEM_HCT:
        required = _HCT_HEADERS
    elif system == SYSTEM_CONTRACT:
        required = _CONTRACT_HEADERS
    elif system == SYSTEM_M00:
        required = _M00_HEADERS
    elif system == SYSTEM_NETSUITE_293:
        required = _NS293_HEADERS
    elif system == SYSTEM_NETSUITE_663:
        required = _NS663_HEADERS
    elif system == SYSTEM_NETSUITE_LOT:
        required = _NS_LOT_HEADERS
    else:
        required = _NS_HEADERS
    match = _find_matching_sheet(book, required)
    if match is None:
        raise InventoryError(
            f"{filename}:找不到包含必要欄位的工作表:" + "、".join(required))
    sheet_name, rows, header_row, header_map = match
    stats.sheet_name = sheet_name

    def col(name: str) -> int | None:
        return header_map.get(name.casefold())

    strip_item_suffix = False
    item_fallback_col = None
    if system == SYSTEM_HCT:
        location_col, item_col = col("儲區類別"), col("客戶產品編號")
        expiry_col, available_col, total_col = col("有效日期"), col("可出數量"), col("庫存數量")
        desc_col = col("產品名稱")
        field_names = ("儲區類別", "客戶產品編號", "有效日期", "可出數量", "庫存數量")
    elif system == SYSTEM_CONTRACT:
        location_col, item_col = col("倉別"), col("品號")
        expiry_col = col("批號")
        available_col = total_col = col("庫存數量")
        desc_col = col("品名")
        field_names = ("倉別", "品號", "批號", "庫存數量", "庫存數量")
    elif system == SYSTEM_M00:
        # 報表沒有倉別欄，一律視為 M00 倉；庫存量＝數量(J)＋組合保留(K)，
        # 兩欄分開解析後於下方相加，同時當作可用量與總庫存量。
        location_col, item_col = None, col("SKU")
        expiry_col = col("有效日期")
        available_col, total_col = col("數量"), col("組合保留")
        desc_col = col("商品名稱")
        field_names = ("倉別", "SKU", "有效日期", "數量", "組合保留")
    elif system == SYSTEM_NETSUITE_293:
        location_col, item_col = col("地點"), col("項目")
        expiry_col = col("到期日")
        available_col = total_col = col("庫存數 總和")
        desc_col = col("項目")
        field_names = ("地點", "項目", "到期日", "庫存數 總和", "庫存數 總和")
        strip_item_suffix = True
    elif system == SYSTEM_NETSUITE_LOT:
        location_col, item_col = col("倉別代碼"), col("料號")
        expiry_col = col("批號")
        available_col = total_col = col("在庫數量")
        desc_col = col("品名")
        field_names = ("倉別代碼", "料號", "批號", "在庫數量", "在庫數量")
    elif system == SYSTEM_NETSUITE_663:
        location_col, item_col = col("地點"), col("DR_料號")
        expiry_col, available_col, total_col = col("庫存編號"), col("可用"), col("在庫量")
        # 2026-07 改版新增「顯示名稱」欄（正式品名），優先於「項目」（料號_序號）
        desc_col = col("顯示名稱")
        if desc_col is None:
            desc_col = col("項目")
        field_names = ("地點", "DR_料號", "庫存編號", "可用", "在庫量")
        item_fallback_col = col("項目")
    else:
        location_col, item_col = col("地點"), col("DR_料號")
        expiry_col, available_col, total_col = col("到期日"), col("項目計數 總和"), col("數量 總和")
        desc_col = col("項目")
        field_names = ("地點", "DR_料號", "到期日", "項目計數 總和", "數量 總和")

    def get(row: list, index: int | None) -> object:
        if index is None or index >= len(row):
            return None
        return row[index]

    for row_index in range(header_row + 1, len(rows)):
        row = rows[row_index]
        raw = [get(row, c) for c in (location_col, item_col, expiry_col, available_col, total_col)]
        if all((v is None or str(v).strip() == "") for v in raw):
            continue

        stats.rows_read += 1
        actual_row = row_index + 1
        has_issue = False

        if system == SYSTEM_CONTRACT and not _normalize_location(get(row, location_col)) \
                and not _normalize_item(get(row, item_col)):
            # 代工廠報表的「XX倉 合計」「總計」小計列：倉別與品號皆空白，列入排除數。
            stats.excluded_rows += 1
            continue

        def anomaly(field_name: str, raw_value: object, issue: str) -> None:
            display_system = SYSTEM_NETSUITE if system.startswith(SYSTEM_NETSUITE) else system
            anomalies.append({
                "來源系統": display_system, "來源檔名": filename, "工作表": sheet_name,
                "原始列號": actual_row, "問題欄位": field_name,
                "原始值": "" if raw_value is None else str(raw_value),
                "異常說明": issue,
            })

        if system == SYSTEM_M00:
            location = M00_LOCATION_PREFIX  # 報表無倉別欄，整份都是 M00 倉
        else:
            location = _normalize_location(get(row, location_col))
        if not location:
            anomaly(field_names[0], get(row, location_col), "倉別不可空白")
            has_issue = True
        elif not location_ok(location):
            stats.excluded_rows += 1
            continue
        elif location_key is not None:
            location = location_key(location)

        item_code = _normalize_item(get(row, item_col))
        if strip_item_suffix and "_" in item_code:
            item_code = item_code.split("_", 1)[0].strip()
        if not item_code and item_fallback_col is not None:
            # 663 格式：DR_料號空白時，改取「項目」底線前段當料號。
            item_code = _normalize_item(get(row, item_fallback_col)).split("_", 1)[0].strip()
        if not item_code:
            anomaly(field_names[1], get(row, item_col), "料號不可空白")
            has_issue = True

        expiry_value = get(row, expiry_col)
        if system == SYSTEM_M00 and str(expiry_value or "").strip() in _M00_NO_EXPIRY_MARKS:
            expiry_value = None  # M00 報表用「-」（或各式破折號）表示無到期日
        expiry_key, expiry_ok = _normalize_date(expiry_value)
        if not expiry_ok:
            anomaly(field_names[2], get(row, expiry_col), "非空日期無法辨識")
            has_issue = True

        available, available_ok = _normalize_quantity(get(row, available_col))
        if not available_ok:
            anomaly(field_names[3], get(row, available_col), "數量不是有效數字")
            has_issue = True

        total, total_ok = _normalize_quantity(get(row, total_col))
        if not total_ok:
            anomaly(field_names[4], get(row, total_col), "數量不是有效數字")
            has_issue = True

        if has_issue:
            stats.anomaly_rows += 1
            continue

        if system == SYSTEM_M00:
            # 單品數量庫存＝數量＋組合保留，同時當作可用量與總庫存量。
            available = total = available + total

        description = _normalize_item(get(row, desc_col)) if desc_col is not None else ""
        _add_aggregate(detail, (location, item_code, expiry_key), available, total, description)
        _add_aggregate(item_totals, (location, item_code), available, total, description)
        stats.valid_rows += 1


def _add_aggregate(aggregate: dict, key: tuple, available: float, total: float,
                   description: str) -> None:
    bucket = aggregate.get(key)
    if bucket is None:
        aggregate[key] = [available, total, 1, description]
    else:
        bucket[0] += available
        bucket[1] += total
        bucket[2] += 1
        if not bucket[3] and description:
            bucket[3] = description


# ------------------------------------------------------------------ 核對


def _qty_equal(a: float, b: float) -> bool:
    """數量比較用容差，避免多筆小數加總的浮點誤差被誤判為差異（同 compare.py）。"""
    return abs(a - b) < 1e-6


def _classify(has_hct: bool, has_ns: bool, hct_avail: float, hct_total: float,
              ns_avail: float, ns_total: float, ext_label: str) -> tuple[str, str]:
    if not has_ns:
        return _ext_only_status(ext_label), f"{ext_label} 有庫存資料，NetSuite 未找到對應資料"
    if not has_hct:
        return STATUS_NS_ONLY, f"NetSuite 有庫存資料，{ext_label} 未找到對應資料"
    if _qty_equal(hct_avail, ns_avail) and _qty_equal(hct_total, ns_total):
        return STATUS_MATCH, f"{ext_label} 與 NetSuite 可用量及總庫存量一致"
    if _qty_equal(hct_total, ns_total):
        return STATUS_AVAILABLE_DIFF, "帳面總量一致，但可出／可用狀態不同"
    if _qty_equal(hct_avail, ns_avail):
        return STATUS_TOTAL_DIFF, "可用量一致，但帳面總量不同"
    return STATUS_BOTH_DIFF, "可用量與總庫存量皆有差異"


def _build_reconciliation(hct_data: dict, ns_data: dict, include_expiry: bool,
                          ext_label: str) -> list[dict]:
    union_keys = sorted(set(hct_data) | set(ns_data), key=lambda k: tuple(str(p).casefold() for p in k))
    result = []
    empty = [0.0, 0.0, 0, ""]
    for key in union_keys:
        has_hct = key in hct_data
        has_ns = key in ns_data
        hct_bucket = hct_data.get(key, empty)
        ns_bucket = ns_data.get(key, empty)
        description = hct_bucket[3] or ns_bucket[3]
        status, explanation = _classify(
            has_hct, has_ns, hct_bucket[0], hct_bucket[1], ns_bucket[0], ns_bucket[1],
            ext_label)
        record = {
            "倉別": key[0],
            "料號": key[1],
            "品名／項目": description,
            f"{ext_label} 可出數量": hct_bucket[0],
            "NetSuite 項目計數": ns_bucket[0],
            "可用量差額": hct_bucket[0] - ns_bucket[0],
            f"{ext_label} 庫存數量": hct_bucket[1],
            "NetSuite 數量": ns_bucket[1],
            "總庫存差額": hct_bucket[1] - ns_bucket[1],
            "狀態": status,
            "結果說明": explanation,
            f"{ext_label} 來源列數": hct_bucket[2],
            "NetSuite 來源列數": ns_bucket[2],
        }
        if include_expiry:
            record["到期日"] = _expiry_output(key[2])
        result.append(record)
    return result


def reconcile(ns_data: bytes, ns_name: str, hct_data: bytes, hct_name: str) -> InventoryResult:
    """核對兩份庫存報表；傳入順序不限，程式自動辨識並對調。"""
    if ns_data == hct_data:
        raise InventoryError("兩次選取的是同一個檔案，請重新選擇不同的來源檔案。")
    first_system = detect_source(ns_data, ns_name)
    if not first_system:
        raise InventoryError(
            f"無法辨識檔案:{ns_name}(找不到 NetSuite、HCT、代工廠或 M00 報表的必要欄位)")
    second_system = detect_source(hct_data, hct_name)
    if not second_system:
        raise InventoryError(
            f"無法辨識檔案:{hct_name}(找不到 NetSuite、HCT、代工廠或 M00 報表的必要欄位)")

    ns_variants = (SYSTEM_NETSUITE, SYSTEM_NETSUITE_293, SYSTEM_NETSUITE_663,
                   SYSTEM_NETSUITE_LOT)
    ext_variants = (SYSTEM_HCT, SYSTEM_CONTRACT, SYSTEM_M00)
    if first_system in ext_variants and second_system in ns_variants:
        ns_data, hct_data = hct_data, ns_data
        ns_name, hct_name = hct_name, ns_name
        ns_system, ext_system = second_system, first_system
    elif first_system in ns_variants and second_system in ext_variants:
        ns_system, ext_system = first_system, second_system
    else:
        raise InventoryError(
            "兩份檔案無法配成一份 NetSuite 與一份 HCT(或代工廠、M00)報表，請重新選取。")

    location_key = None
    if ext_system == SYSTEM_M00:
        # M00 對帳只看 M00 開頭倉別(NetSuite 側；M00 報表本身整份視為 M00 倉)，
        # 並把通過過濾的倉別一律收斂成「M00」當核對鍵。
        def location_ok(location: str) -> bool:
            return location.startswith(M00_LOCATION_PREFIX)

        def location_key(_location: str) -> str:
            return M00_LOCATION_PREFIX
        ext_label = SYSTEM_M00
        scope_text = f"{M00_LOCATION_PREFIX} 開頭電商物流倉"
        excluded_label = "排除非 M00 倉筆數"
        output_prefix = "M00庫存核對結果"
    elif ext_system == SYSTEM_CONTRACT:
        # 代工廠對帳只看 D 開頭代工廠倉(兩邊都套同一條規則)。
        def location_ok(location: str) -> bool:
            return location.startswith(CONTRACT_LOCATION_PREFIX)
        ext_label = SYSTEM_CONTRACT
        scope_text = f"{CONTRACT_LOCATION_PREFIX} 開頭代工廠倉"
        excluded_label = "排除非 D 倉／合計列筆數"
        output_prefix = "代工廠庫存核對結果"
    else:
        def location_ok(location: str) -> bool:
            return location in APPROVED_LOCATIONS
        ext_label = SYSTEM_HCT
        scope_text = "、".join(sorted(APPROVED_LOCATIONS))
        excluded_label = "排除非 G 倉筆數"
        output_prefix = "庫存核對結果"

    hct_detail: dict = {}
    hct_items: dict = {}
    ns_detail: dict = {}
    ns_items: dict = {}
    anomalies: list = []
    hct_stats = SourceStats(file_name=hct_name)
    ns_stats = SourceStats(file_name=ns_name)

    _import_source(ns_data, ns_name, ns_system, ns_detail, ns_items, anomalies, ns_stats,
                   location_ok, location_key)
    _import_source(hct_data, hct_name, ext_system, hct_detail, hct_items, anomalies, hct_stats,
                   location_ok, location_key)

    detail_rows = _build_reconciliation(hct_detail, ns_detail, include_expiry=True,
                                        ext_label=ext_label)
    item_rows = _build_reconciliation(hct_items, ns_items, include_expiry=False,
                                      ext_label=ext_label)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        output_bytes = _write_output(detail_rows, item_rows, anomalies, hct_stats, ns_stats,
                                     ext_label, scope_text, excluded_label)
    except Exception as exc:
        raise InventoryError(f"建立輸出檔時發生錯誤：{exc}") from exc

    def counts(rows):
        counter: dict[str, int] = {}
        for record in rows:
            counter[record["狀態"]] = counter.get(record["狀態"], 0) + 1
        return counter

    return InventoryResult(
        output_bytes=output_bytes,
        output_name=f"{output_prefix}_{stamp}.xlsx",
        detail_rows=detail_rows,
        item_rows=item_rows,
        anomalies=anomalies,
        hct_stats=hct_stats,
        ns_stats=ns_stats,
        detail_status_counts=counts(detail_rows),
        item_status_counts=counts(item_rows),
        ext_label=ext_label,
        statuses=status_order(ext_label),
    )


# ------------------------------------------------------------------ 輸出

def _detail_columns(ext_label: str) -> list[str]:
    return [
        "倉別", "料號", "品名／項目", "到期日", f"{ext_label} 可出數量", "NetSuite 項目計數",
        "可用量差額", f"{ext_label} 庫存數量", "NetSuite 數量", "總庫存差額", "狀態", "結果說明",
        f"{ext_label} 來源列數", "NetSuite 來源列數",
    ]


def _item_columns(ext_label: str) -> list[str]:
    columns = _detail_columns(ext_label)
    columns.remove("到期日")
    return columns


_ANOMALY_COLUMNS = ["來源系統", "來源檔名", "工作表", "原始列號", "問題欄位", "原始值", "異常說明"]


def _status_fill(ext_label: str) -> dict:
    return {
        STATUS_MATCH: ("C6EFCE", "006100"),
        STATUS_AVAILABLE_DIFF: ("FFEB9C", "9C6500"),
        STATUS_TOTAL_DIFF: ("FFEB9C", "9C6500"),
        STATUS_BOTH_DIFF: ("FFC7CE", "9C0006"),
        _ext_only_status(ext_label): ("FFC7CE", "9C0006"),
        STATUS_NS_ONLY: ("FFC7CE", "9C0006"),
    }


def _write_output(detail_rows, item_rows, anomalies, hct_stats, ns_stats,
                  ext_label=SYSTEM_HCT, scope_text="", excluded_label="排除非 G 倉筆數") -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = "核對摘要"
    summary.merge_cells("A1:F1")
    summary["A1"] = "庫存核對摘要"
    summary["A1"].font = Font(bold=True, size=20, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    summary.row_dimensions[1].height = 38

    summary["A3"] = "狀態"
    summary["B3"] = "日期明細"
    summary["C3"] = "料號彙總"

    def count_rows(rows, status):
        return sum(1 for r in rows if r["狀態"] == status)

    for offset, status in enumerate(status_order(ext_label)):
        summary.cell(4 + offset, 1, status)
        summary.cell(4 + offset, 2, count_rows(detail_rows, status))
        summary.cell(4 + offset, 3, count_rows(item_rows, status))
    summary["A10"] = "總筆數"
    summary["B10"] = len(detail_rows)
    summary["C10"] = len(item_rows)

    summary["A12"] = "核對範圍倉別"
    summary["B12"] = scope_text or "、".join(sorted(APPROVED_LOCATIONS))
    summary["A13"] = "資料統計"
    summary["A14"] = "項目"
    summary["B14"] = ext_label
    summary["C14"] = "NetSuite"
    stat_labels = [
        ("讀取筆數", "rows_read"), ("有效筆數", "valid_rows"),
        (excluded_label, "excluded_rows"), ("異常來源列數", "anomaly_rows"),
    ]
    for offset, (label, attr) in enumerate(stat_labels):
        summary.cell(15 + offset, 1, label)
        summary.cell(15 + offset, 2, getattr(hct_stats, attr))
        summary.cell(15 + offset, 3, getattr(ns_stats, attr))
    summary["A19"] = "異常明細筆數"
    summary["B19"] = len(anomalies)
    summary["A20"] = f"{ext_label} 檔案"
    summary["B20"] = hct_stats.file_name
    summary["A21"] = "NetSuite 檔案"
    summary["B21"] = ns_stats.file_name
    summary["A22"] = "核對時間"
    summary["B22"] = datetime.now()
    summary["B22"].number_format = "yyyy-mm-dd hh:mm:ss"

    light_fill = PatternFill("solid", fgColor="D9E1F2")
    for ref in ("A3", "B3", "C3", "A13", "A14", "B14", "C14"):
        summary[ref].font = Font(bold=True)
        summary[ref].fill = light_fill
    summary["A4"].fill = PatternFill("solid", fgColor="C6EFCE")
    for ref in ("A5", "A6"):
        summary[ref].fill = PatternFill("solid", fgColor="FFEB9C")
    for ref in ("A7", "A8", "A9"):
        summary[ref].fill = PatternFill("solid", fgColor="FFC7CE")
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 40
    summary.column_dimensions["C"].width = 18

    def write_table(title: str, columns: list, rows: list, status_col: int | None):
        ws = wb.create_sheet(title)
        for col, header in enumerate(columns, start=1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 34
        for row_offset, record in enumerate(rows, start=2):
            for col, header in enumerate(columns, start=1):
                value = record.get(header, "")
                cell = ws.cell(row_offset, col, value)
                if header == "到期日" and isinstance(value, date):
                    cell.number_format = "yyyy-mm-dd"
                elif header == "料號":
                    cell.number_format = "@"
                elif isinstance(value, (int, float)) and header not in ("原始列號",):
                    cell.number_format = "#,##0.####"
                if status_col is not None and col == status_col:
                    colors = _status_fill(ext_label).get(str(value))
                    if colors:
                        cell.fill = PatternFill("solid", fgColor=colors[0])
                        cell.font = Font(color=colors[1])
        if not rows:
            ws.cell(2, 1, "無資料" if status_col is not None else "無資料異常")
        import openpyxl.utils
        last_col = openpyxl.utils.get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{max(len(rows) + 1, 1)}"
        ws.freeze_panes = "A2"
        widths = {"倉別": 10, "料號": 18, "品名／項目": 34, "到期日": 14,
                  "狀態": 18, "結果說明": 42, "異常說明": 32, "原始值": 32, "來源檔名": 28}
        for col, header in enumerate(columns, start=1):
            letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[letter].width = widths.get(header, 16)

    write_table("日期明細", _detail_columns(ext_label), detail_rows, status_col=11)
    write_table("料號彙總", _item_columns(ext_label), item_rows, status_col=10)
    write_table("資料異常", _ANOMALY_COLUMNS, anomalies, status_col=None)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
