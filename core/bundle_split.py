# -*- coding: utf-8 -*-
"""ED 訂單明細組合品拆解：把 299 開頭的「組合料號」展開成單品明細。

NetSuite「每日銷售串接訂單明細」裡，架上組（虛擬組合套件，料號多為 299 開頭）
是一列一個套件，倉庫端與庫存端看到的卻是單品。本模組依組合對照表把套件列
展開成單品列，供後續串接／核對使用。

拆解規則（與 Shopline 日結工具 core/bom.py 一致）：
  - 數量　　＝ 套件數量 × 每套數量
  - 含稅金額＝ 依分攤比率拆分，四捨五入的尾差併入比率最大的單品，
    確保拆解前後金額總計完全一致（來源金額是整數時，拆出來也維持整數）
  - 結帳單價＝ 拆分後金額 ÷ 拆分後數量（以金額為準回推）
  - 其餘欄位（單號、客戶、發票資訊…）原樣複製到每一列

對照表支援兩種格式，載入時自動判別：
  - 簡化（可編輯）格式：套件品號／單品品號／單品品名／規格／每套數量／分攤比率
  - ERP 原始匯出格式：MDxxx_中文 欄名，同一套件有多版時取「生效日期最新」那版
分攤比率一律以套件為單位正規化（加總＝1）；比率缺漏時改用標準售價×數量加權，
再不行就平均分攤。
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime

from .xlio import (
    build_header_map,
    first_sheet,
    is_blank_row,
    normalize_number,
    normalize_text,
    read_workbook,
)

# ---------------------------------------------------------------- 對照表欄名

SIMPLE_COL_BUNDLE = "套件品號"
SIMPLE_COL_SKU = "單品品號"
SIMPLE_COL_NAME = "單品品名"
SIMPLE_COL_SPEC = "規格"
SIMPLE_COL_QTY = "每套數量"
SIMPLE_COL_RATIO = "分攤比率"

ERP_COL_BUNDLE = "MD003_組合套件品號"
ERP_COL_CUSTOMER = "MD001_客戶代號"
ERP_COL_EFFECTIVE = "MD005_生效日期"
ERP_COL_SKU = "MD006_品號"
ERP_COL_NAME = "MD007_品名"
ERP_COL_SPEC = "MD010_規格"
ERP_COL_QTY = "MD014_數量"
ERP_COL_PRICE = "MD016_標準售價"
ERP_COL_RATIO = "MD018_分攤比率"

# ---------------------------------------------------------------- 明細欄名別名

_SKU_ALIASES = ("商品貨號", "料號", "DR_料號", "品號", "商品編號", "產品編號", "存貨編號")
_NAME_ALIASES = ("商品名稱", "品名", "商品名", "顯示名稱", "產品名稱")
_QTY_ALIASES = ("數量", "商品數量", "銷售數量")
_AMOUNT_ALIASES = ("含稅金額", "金額", "含稅小計", "銷售金額", "小計")
_PRICE_ALIASES = ("結帳單價", "單價", "售價", "成交單價")

# 不在對照表、但料號是 299 開頭（架上組虛擬組合）時提出提醒，避免該拆的漏拆。
# 只是提醒，不影響輸出：這類列會原樣保留。
# 298 開頭是正常組合品（ERP 有自己的品號與庫存），本來就不拆，不列入提醒。
_BUNDLE_SKU_PREFIXES = ("299",)

# 拆解後附加在原欄位右邊的稽核欄位
EXTRA_COLUMNS = ["拆解狀態", "組合來源料號", "組合來源品名", "每套數量", "分攤比率"]
STATUS_KEPT = "原樣保留"
STATUS_SPLIT = "組合拆解"

# 介面上只預覽前幾列（完整結果請下載 xlsx），避免大檔案把 session 撐大
PREVIEW_LIMIT = 300


class BundleSplitError(ValueError):
    pass


@dataclass
class BundleComponent:
    sku: str
    name: str
    spec: str
    qty: float    # 每 1 套所含數量
    ratio: float  # 金額分攤比率（正規化後每套件加總＝1）


@dataclass
class BundleSplitResult:
    output_bytes: bytes
    output_name: str
    source_name: str
    map_name: str
    bundle_count: int          # 對照表載入的套件數
    source_rows: int           # 來源明細列數
    output_rows: int           # 拆解後明細列數
    split_rows: int            # 被拆解掉的套件列數
    component_rows: int        # 拆解產生的單品列數
    used_bundles: int          # 實際用到的套件品號數
    amount_before: float
    amount_after: float
    unmapped_rows: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    preview_columns: list = field(default_factory=list)
    preview_rows: list = field(default_factory=list)

    @property
    def amount_matched(self) -> bool:
        return abs(self.amount_before - self.amount_after) < 0.005


# ------------------------------------------------------------------ 對照表載入


def load_bundle_map(data: bytes, filename: str = "") -> dict[str, list[BundleComponent]]:
    """讀組合對照表（.xls/.xlsx/SpreadsheetML 皆可），回傳 {套件品號: [BundleComponent]}。"""
    rows = first_sheet(read_workbook(data, filename))
    headers = build_header_map(rows[0]) if rows else {}
    body = [row for row in rows[1:] if not is_blank_row(row)]
    if ERP_COL_BUNDLE.casefold() in headers:
        bundle_map = _load_erp(body, headers)
    elif SIMPLE_COL_BUNDLE.casefold() in headers:
        bundle_map = _load_simple(body, headers)
    else:
        raise BundleSplitError(
            f"無法辨識組合對照表的欄位：{filename or '（未命名檔案）'}\n"
            f"可編輯格式需要欄位：{SIMPLE_COL_BUNDLE}、{SIMPLE_COL_SKU}、{SIMPLE_COL_QTY}、{SIMPLE_COL_RATIO}\n"
            f"ERP 原始格式需要欄位：{ERP_COL_BUNDLE}、{ERP_COL_SKU}、{ERP_COL_QTY}、{ERP_COL_RATIO}"
        )
    if not bundle_map:
        raise BundleSplitError(f"組合對照表沒有任何有效資料列：{filename or '（未命名檔案）'}")
    return bundle_map


def _cell(row: list, headers: dict[str, int], name: str) -> object:
    index = headers.get(name.casefold())
    if index is None or index >= len(row):
        return None
    return row[index]


def _load_simple(body: list, headers: dict[str, int]) -> dict[str, list[BundleComponent]]:
    out: dict[str, list[BundleComponent]] = {}
    for row in body:
        bundle = normalize_text(_cell(row, headers, SIMPLE_COL_BUNDLE))
        sku = normalize_text(_cell(row, headers, SIMPLE_COL_SKU))
        if not bundle or not sku:
            continue
        out.setdefault(bundle, []).append(
            BundleComponent(
                sku=sku,
                name=normalize_text(_cell(row, headers, SIMPLE_COL_NAME)),
                spec=normalize_text(_cell(row, headers, SIMPLE_COL_SPEC)),
                qty=normalize_number(_cell(row, headers, SIMPLE_COL_QTY)) or 1.0,
                ratio=normalize_number(_cell(row, headers, SIMPLE_COL_RATIO)),
            )
        )
    for comps in out.values():
        _normalize_ratios(comps)
    return out


def _load_erp(body: list, headers: dict[str, int]) -> dict[str, list[BundleComponent]]:
    """ERP 原始格式：同一套件常在多個客戶代號／生效日期下各有一版，取生效日期最新的。"""
    versions: dict[str, dict[tuple, list[list]]] = {}
    for row in body:
        bundle = normalize_text(_cell(row, headers, ERP_COL_BUNDLE))
        sku = normalize_text(_cell(row, headers, ERP_COL_SKU))
        if not bundle or not sku:
            continue
        version_key = (
            normalize_text(_cell(row, headers, ERP_COL_CUSTOMER)),
            normalize_text(_cell(row, headers, ERP_COL_EFFECTIVE)),
        )
        versions.setdefault(bundle, {}).setdefault(version_key, []).append(row)

    out: dict[str, list[BundleComponent]] = {}
    for bundle, vers in versions.items():
        version_key = max(vers, key=lambda k: (k[1], len(vers[k])))  # 生效日期最新，同日取列數多者
        rows = vers[version_key]
        comps = [
            BundleComponent(
                sku=normalize_text(_cell(row, headers, ERP_COL_SKU)),
                name=normalize_text(_cell(row, headers, ERP_COL_NAME)),
                spec=normalize_text(_cell(row, headers, ERP_COL_SPEC)),
                qty=normalize_number(_cell(row, headers, ERP_COL_QTY)) or 1.0,
                ratio=normalize_number(_cell(row, headers, ERP_COL_RATIO)),
            )
            for row in rows
        ]
        weights = [
            normalize_number(_cell(row, headers, ERP_COL_PRICE))
            * (normalize_number(_cell(row, headers, ERP_COL_QTY)) or 1.0)
            for row in rows
        ]
        _normalize_ratios(comps, weights)
        out[bundle] = comps
    return out


def _normalize_ratios(comps: list[BundleComponent], weights: list[float] | None = None) -> None:
    """把分攤比率正規化成加總＝1；比率缺漏時改用 weights（售價×數量），再不行平均分攤。"""
    if not comps:
        return
    total = sum(c.ratio for c in comps)
    if total > 0:
        for c in comps:
            c.ratio = c.ratio / total
        return
    weight_total = sum(weights) if weights else 0.0
    if weights and weight_total > 0:
        for c, w in zip(comps, weights):
            c.ratio = w / weight_total
        return
    for c in comps:
        c.ratio = 1.0 / len(comps)


# ------------------------------------------------------------------ 金額分攤


def _amount_digits(gross: float) -> int:
    """來源金額是整數（台幣明細常見）就拆成整數，有小數才拆到 2 位。"""
    return 0 if abs(gross - round(gross)) < 1e-9 else 2


def allocate(gross: float, ratios: list[float], digits: int = 2) -> list[float]:
    """把含稅金額依比率拆給各單品；四捨五入尾差併入比率最大者，總計不變。"""
    if not ratios:
        return []
    amounts = [round(gross * r, digits) for r in ratios]
    diff = round(gross - sum(amounts), digits)
    if diff:
        index = max(range(len(ratios)), key=lambda j: ratios[j])
        amounts[index] = round(amounts[index] + diff, digits)
    if digits == 0:
        return [float(int(a)) for a in amounts]
    return amounts


def _clean_number(value: float) -> float | int:
    """整數值輸出成 int，避免 Excel 出現 10.0 這種看起來像小數的數量。"""
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return value


# ------------------------------------------------------------------ 拆解


def split(
    order_data: bytes,
    order_name: str,
    map_data: bytes,
    map_name: str,
) -> BundleSplitResult:
    """讀訂單明細與組合對照表，回傳拆解結果（含可下載的 xlsx）。"""
    bundle_map = load_bundle_map(map_data, map_name)
    rows = first_sheet(read_workbook(order_data, order_name))
    return split_rows(rows, bundle_map, order_name, map_name)


def split_rows(
    rows: list[list[object]],
    bundle_map: dict[str, list[BundleComponent]],
    source_name: str = "",
    map_name: str = "",
) -> BundleSplitResult:
    """對已讀成 2D list 的訂單明細做拆解（第一列為表頭）。"""
    if len(rows) < 2:
        raise BundleSplitError("訂單明細沒有任何資料列（只有表頭或整份空白）。")

    header = [normalize_text(value) for value in rows[0]]
    headers = build_header_map(rows[0])
    warnings: list[str] = []

    sku_index = _resolve_column(headers, _SKU_ALIASES)
    if sku_index is None:
        raise BundleSplitError(
            "訂單明細找不到商品貨號欄位。\n"
            f"可接受的欄名：{'、'.join(_SKU_ALIASES)}"
        )
    qty_index = _resolve_column(headers, _QTY_ALIASES)
    if qty_index is None:
        raise BundleSplitError(
            "訂單明細找不到數量欄位。\n"
            f"可接受的欄名：{'、'.join(_QTY_ALIASES)}"
        )
    name_index = _resolve_column(headers, _NAME_ALIASES)
    amount_index = _resolve_column(headers, _AMOUNT_ALIASES)
    price_index = _resolve_column(headers, _PRICE_ALIASES)
    if amount_index is None:
        warnings.append(
            "找不到含稅金額欄位，拆解後只展開數量、不分攤金額。"
            f"（可接受的欄名：{'、'.join(_AMOUNT_ALIASES)}）"
        )
    if name_index is None:
        warnings.append("找不到商品名稱欄位，拆解後不會替換成單品品名。")

    body = [row for row in rows[1:] if not is_blank_row(row)]
    width = max([len(header)] + [len(row) for row in body])
    header = header + [""] * (width - len(header))

    out_rows: list[list[object]] = []
    split_rows_count = 0
    component_rows = 0
    used_bundles: set[str] = set()
    unmapped: dict[str, dict] = {}
    amount_before = 0.0
    amount_after = 0.0

    for row in body:
        padded = list(row) + [None] * (width - len(row))
        sku = normalize_text(padded[sku_index])
        gross = normalize_number(padded[amount_index]) if amount_index is not None else 0.0
        amount_before += gross

        comps = bundle_map.get(sku)
        if not comps:
            if sku and sku.startswith(_BUNDLE_SKU_PREFIXES):
                record = unmapped.setdefault(sku, {
                    "商品貨號": sku,
                    "商品名稱": normalize_text(padded[name_index]) if name_index is not None else "",
                    "出現列數": 0,
                    "數量合計": 0.0,
                    "含稅金額合計": 0.0,
                })
                record["出現列數"] += 1
                record["數量合計"] += normalize_number(padded[qty_index])
                record["含稅金額合計"] += gross
            out_rows.append(padded + [STATUS_KEPT, "", "", "", ""])
            amount_after += gross
            continue

        split_rows_count += 1
        used_bundles.add(sku)
        qty = normalize_number(padded[qty_index])
        bundle_name = normalize_text(padded[name_index]) if name_index is not None else ""
        amounts = (
            allocate(gross, [c.ratio for c in comps], _amount_digits(gross))
            if amount_index is not None else [0.0] * len(comps)
        )
        for comp, amount in zip(comps, amounts):
            new_row = list(padded)
            new_row[sku_index] = comp.sku
            if name_index is not None:
                new_row[name_index] = comp.name or bundle_name
            comp_qty = qty * comp.qty
            new_row[qty_index] = _clean_number(comp_qty)
            if amount_index is not None:
                new_row[amount_index] = _clean_number(amount)
                amount_after += amount
            if price_index is not None and amount_index is not None:
                new_row[price_index] = round(amount / comp_qty, 4) if comp_qty else 0
            out_rows.append(new_row + [
                STATUS_SPLIT,
                sku,
                bundle_name,
                _clean_number(comp.qty),
                round(comp.ratio, 6),
            ])
            component_rows += 1

    if unmapped:
        warnings.append(
            f"有 {len(unmapped)} 個料號看起來是組合品（{'／'.join(_BUNDLE_SKU_PREFIXES)} 開頭）"
            "但不在對照表裡，已原樣保留未拆解，詳見輸出檔的「未對應組合料號」工作表。"
        )

    unmapped_rows = sorted(unmapped.values(), key=lambda r: r["商品貨號"])
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result = BundleSplitResult(
        output_bytes=b"",
        output_name=f"ED訂單明細拆解結果_{stamp}.xlsx",
        source_name=source_name,
        map_name=map_name,
        bundle_count=len(bundle_map),
        source_rows=len(body),
        output_rows=len(out_rows),
        split_rows=split_rows_count,
        component_rows=component_rows,
        used_bundles=len(used_bundles),
        amount_before=round(amount_before, 2),
        amount_after=round(amount_after, 2),
        unmapped_rows=unmapped_rows,
        warnings=warnings,
        preview_columns=header + EXTRA_COLUMNS,
        preview_rows=[list(row) for row in out_rows[:PREVIEW_LIMIT]],
    )
    if not result.amount_matched:
        result.warnings.append(
            f"拆解前後含稅金額不一致（前 {result.amount_before}、後 {result.amount_after}），請回報此問題。"
        )
    try:
        result.output_bytes = _write_output(header + EXTRA_COLUMNS, out_rows, result)
    except Exception as exc:  # noqa: BLE001 - 轉成使用者看得懂的訊息
        raise BundleSplitError(f"建立輸出檔時發生錯誤：{exc}") from exc
    return result


def _resolve_column(headers: dict[str, int], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        index = headers.get(alias.casefold())
        if index is not None:
            return index
    return None


# ------------------------------------------------------------------ 輸出

_UNMAPPED_COLUMNS = ["商品貨號", "商品名稱", "出現列數", "數量合計", "含稅金額合計"]


def _write_output(header: list, out_rows: list, result: BundleSplitResult) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    split_fill = PatternFill("solid", fgColor="E2EFDA")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = "拆解摘要"
    summary["A1"] = "ED 訂單明細組合品拆解摘要"
    summary["A1"].font = Font(bold=True, size=18)
    info = [
        ("產生時間", datetime.now()),
        ("來源訂單明細", result.source_name),
        ("組合對照表", result.map_name),
        ("對照表套件數", result.bundle_count),
        ("來源明細列數", result.source_rows),
        ("拆解後明細列數", result.output_rows),
        ("被拆解的組合列數", result.split_rows),
        ("拆解產生的單品列數", result.component_rows),
        ("實際用到的套件品號數", result.used_bundles),
        ("拆解前含稅金額合計", result.amount_before),
        ("拆解後含稅金額合計", result.amount_after),
        ("金額一致性", "一致" if result.amount_matched else "不一致（請回報）"),
        ("未對應組合料號數", len(result.unmapped_rows)),
    ]
    for offset, (label, value) in enumerate(info, start=3):
        summary.cell(offset, 1, label).font = Font(bold=True)
        cell = summary.cell(offset, 2, value)
        if isinstance(value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 52

    ws = wb.create_sheet("拆解後明細")
    for col, title in enumerate(header, start=1):
        cell = ws.cell(1, col, title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    status_col = len(header) - len(EXTRA_COLUMNS) + 1
    for row_offset, values in enumerate(out_rows, start=2):
        is_split = values[status_col - 1] == STATUS_SPLIT
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_offset, col, value)
            cell.border = border
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(value, date):
                cell.number_format = "yyyy-mm-dd"
            if is_split:
                cell.fill = split_fill
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{max(len(out_rows) + 1, 1)}"
    ws.freeze_panes = "A2"
    _autosize(ws, header, [list(r) for r in out_rows])

    if result.unmapped_rows:
        ws2 = wb.create_sheet("未對應組合料號")
        for col, title in enumerate(_UNMAPPED_COLUMNS, start=1):
            cell = ws2.cell(1, col, title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for row_offset, record in enumerate(result.unmapped_rows, start=2):
            for col, title in enumerate(_UNMAPPED_COLUMNS, start=1):
                cell = ws2.cell(row_offset, col, _clean_number(record[title])
                                if isinstance(record[title], float) else record[title])
                cell.border = border
        ws2.freeze_panes = "A2"
        _autosize(ws2, _UNMAPPED_COLUMNS,
                  [[r[c] for c in _UNMAPPED_COLUMNS] for r in result.unmapped_rows])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _autosize(ws, header: list, rows: list[list]) -> None:
    from openpyxl.utils import get_column_letter

    for col in range(1, len(header) + 1):
        widths = [len(str(header[col - 1])) * 2]
        for row in rows[:200]:
            if col - 1 < len(row) and row[col - 1] is not None:
                widths.append(min(len(str(row[col - 1])), 40))
        ws.column_dimensions[get_column_letter(col)].width = max(10, min(max(widths) + 2, 44))
