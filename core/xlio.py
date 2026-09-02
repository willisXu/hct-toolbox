# -*- coding: utf-8 -*-
"""萬用 Excel 讀檔器。

NetSuite 匯出的 .xls 其實是 SpreadsheetML XML；HCT 系統匯出的 .xls 是傳統 BIFF；
使用者也可能丟 .xlsx / .xlsm。本模組把三種格式都讀成
{工作表名稱: 2D list(每列一個 list，值為 str/float/datetime/None)}。
"""
from __future__ import annotations

import io
import json
import re
import xml.etree.ElementTree as ET
from datetime import datetime, date
from pathlib import Path


SSNS = "{urn:schemas-microsoft-com:office:spreadsheet}"


def read_workbook(data: bytes, filename: str = "") -> dict[str, list[list[object]]]:
    """讀取 Excel 檔案內容，回傳 {sheet_name: rows}。"""
    if data[:3] == b"\xef\xbb\xbf":
        # UTF-8 BOM 會讓 XML 判斷失敗（ET.fromstring 也不接受 bytes 前的 BOM）
        data = data[3:]
    head = data[:512].lstrip()
    if head.startswith(b"<?xml") or head.startswith(b"<Workbook"):
        return _read_spreadsheetml(data, filename)
    if head[:100].lower().startswith((b"<html", b"<!doctype")):
        raise ValueError(f"不支援 HTML 格式的報表：{filename}")
    if data[:4] == b"\xd0\xcf\x11\xe0":
        return _read_biff(data, filename)
    if data[:2] == b"PK":
        return _read_openxml(data, filename)
    raise ValueError(f"無法辨識的檔案格式：{filename}")


def _read_spreadsheetml(data: bytes, filename: str = "") -> dict[str, list[list[object]]]:
    try:
        root = ET.fromstring(data)
    except ET.ParseError as exc:
        raise ValueError(f"讀取檔案失敗：{filename}\n檔案內容已毀損或不是有效的 XML 報表（{exc}）") from exc
    result: dict[str, list[list[object]]] = {}
    for ws in root.findall(f"{SSNS}Worksheet"):
        name = ws.get(f"{SSNS}Name") or f"Sheet{len(result) + 1}"
        table = ws.find(f"{SSNS}Table")
        rows: list[list[object]] = []
        if table is not None:
            for row in table.findall(f"{SSNS}Row"):
                values: list[object] = []
                for cell in row.findall(f"{SSNS}Cell"):
                    index_text = cell.get(f"{SSNS}Index")
                    if index_text:
                        while len(values) < int(index_text) - 1:
                            values.append(None)
                    data_el = cell.find(f"{SSNS}Data")
                    values.append(_convert_ss_value(data_el))
                rows.append(values)
        result[name] = _pad_rows(rows)
    return result


def _convert_ss_value(data_el: ET.Element | None) -> object:
    if data_el is None:
        return None
    text = "".join(data_el.itertext())
    value_type = data_el.get(f"{SSNS}Type", "String")
    if value_type == "Number":
        try:
            return float(text)
        except ValueError:
            return text
    if value_type == "DateTime":
        try:
            return datetime.fromisoformat(text.rstrip("Z"))
        except ValueError:
            return text
    if value_type == "Boolean":
        return text == "1"
    return text


def _read_biff(data: bytes, filename: str = "") -> dict[str, list[list[object]]]:
    import xlrd

    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as exc:
        raise ValueError(
            f"讀取檔案失敗：{filename}\n檔案內容已毀損或不是有效的 Excel 97-2003 報表（{exc}）"
        ) from exc
    result: dict[str, list[list[object]]] = {}
    for sheet in book.sheets():
        rows: list[list[object]] = []
        for r in range(sheet.nrows):
            values: list[object] = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    values.append(datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode)))
                elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    values.append(None)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    values.append(bool(cell.value))
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    values.append(None)
                else:
                    values.append(cell.value)
            rows.append(values)
        result[sheet.name] = _pad_rows(rows)
    return result


def _read_openxml(data: bytes, filename: str = "") -> dict[str, list[list[object]]]:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(
            f"讀取檔案失敗：{filename}\n檔案內容已毀損或不是有效的 Excel 報表（{exc}）"
        ) from exc
    result: dict[str, list[list[object]]] = {}
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        result[ws.title] = _pad_rows(rows)
    wb.close()
    return result


def _pad_rows(rows: list[list[object]]) -> list[list[object]]:
    width = max((len(r) for r in rows), default=0)
    for r in rows:
        while len(r) < width:
            r.append(None)
    return rows


def first_sheet(book: dict[str, list[list[object]]]) -> list[list[object]]:
    """回傳第一個工作表（比照 VBA 只讀取 Worksheets(1) 的行為，不會自動改讀其他工作表）。"""
    if not book:
        raise ValueError("檔案中沒有可讀取的表格資料。")
    rows = next(iter(book.values()))
    if not rows:
        raise ValueError("檔案中沒有可讀取的表格資料。")
    return rows


# ---------------------------------------------------------------- 共用正規化

_EXCEL_ERROR_VALUES = {
    "#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#GETTING_DATA",
}


def is_excel_error_text(value: object) -> bool:
    """判斷是否為 Excel 公式錯誤值的文字表示（VBA 端用 IsError 偵測）。"""
    return isinstance(value, str) and value.strip().upper() in _EXCEL_ERROR_VALUES


def build_header_map(header_row: list) -> dict[str, int]:
    """欄名（casefold）→ 欄索引；同名欄取第一個。各模組共用，避免各自維護。"""
    headers: dict[str, int] = {}
    for index, value in enumerate(header_row):
        text = normalize_text(value)
        if text and text.casefold() not in headers:
            headers[text.casefold()] = index
    return headers


# NetSuite saved search 的欄位沒設中文自訂 Label 時，匯出檔／RESTlet 拿到的是
# 該欄位的英文預設名稱（Document Number、Memo (Main)…），跟轉換邏輯認的中文
# 欄名對不上。同一份報表常常一半中文一半英文——有設 Label 的欄（來源倉別、
# 出貨數量…）是中文，沒設的就退回英文，所以不能只靠「整份是中文還是英文」判斷。
#
# 對照表左邊同時收兩種寫法：
#   1. 英文預設名稱（手動匯出檔、RESTlet 的 col.label 都是這個）
#   2. 欄位內部 ID（連 Label 都沒有時 RESTlet 取 col.name 回傳 displayname 之類）
NETSUITE_HEADER_ALIASES = {
    # 英文預設名稱
    "document number": "文件編號",
    "date": "日期",
    "internal id": "內部 ID",
    "item": "項目",
    "display name": "顯示名稱",
    "description": "說明",
    "memo": "備忘錄",
    "memo (main)": "備忘錄 (主要)",
    "transaction serial/lot number": "交易序號/批號",
    "serial/lot number": "交易序號/批號",
    "transaction line type": "交易明細行類型",
    "quantity": "數量",
    "location": "地點",
    # 欄位內部 ID（saved search 沒設 Label 時回傳的值）
    "tranid": "文件編號",
    "trandate": "日期",
    "internalid": "內部 ID",
    "itemid": "項目",
    "displayname": "顯示名稱",
    "salesdescription": "說明",
    "memomain": "備忘錄 (主要)",
}


# 從 NetSuite 自動學回來的欄名對照（見 core/netsuite.py learn_header_aliases）。
# 內建表是人工驗證過的少數常見欄名，這份則是掃過帳號裡的 saved search、用
# 「同一個欄位內部 ID 在別支 saved search 有中文 Label」推出來的，涵蓋公司自己
# 那些欄。存成檔案是為了讓「上傳檔案」路徑（根本沒連 NetSuite）也用得到。
HEADER_ALIAS_CACHE_PATH = Path(__file__).resolve().parent.parent / "mappings" / "欄位對照快取.json"

_CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
_learned_cache: dict | None = None


def has_cjk(text: object) -> bool:
    """含中日韓文字＝這是人設過的中文 Label，不是 NetSuite 的英文預設名稱。"""
    return bool(_CJK_RE.search(str(text or "")))


def load_header_alias_cache(refresh: bool = False) -> dict:
    """讀自動對照快取；檔案不存在／壞掉都當成空的（這只是加分項，不能擋住轉換）。"""
    global _learned_cache
    if _learned_cache is not None and not refresh:
        return _learned_cache
    try:
        data = json.loads(HEADER_ALIAS_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        data = {}
    if not isinstance(data, dict):
        data = {}
    aliases = data.get("aliases")
    data["aliases"] = {
        str(k).casefold(): str(v)
        for k, v in (aliases or {}).items()
        if str(k).strip() and str(v).strip()
    }
    _learned_cache = data
    return data


def save_header_alias_cache(aliases: dict[str, str], sources: list[dict]) -> Path:
    """把學到的對照寫成檔案，並讓後續呼叫立刻吃到新的內容。"""
    global _learned_cache
    payload = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "aliases": {str(k).casefold(): str(v) for k, v in aliases.items()},
        "sources": sources,
    }
    HEADER_ALIAS_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    HEADER_ALIAS_CACHE_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _learned_cache = payload
    return HEADER_ALIAS_CACHE_PATH


def header_aliases() -> dict[str, str]:
    """內建表 + 自動學回來的對照。衝突時以內建表為準（人工驗證過的優先）。"""
    merged = dict(load_header_alias_cache().get("aliases") or {})
    merged.update(NETSUITE_HEADER_ALIASES)
    return merged


def apply_netsuite_header_aliases(headers: dict[str, int]) -> None:
    """把 header map 裡的英文預設欄名／欄位內部 ID 補上對應的中文欄名（就地修改）。

    只補、不覆蓋：來源檔本來就有該中文欄時以中文欄為準（實際匯出檔會同時出現
    「內部 ID」與「Internal ID」兩欄），英文鍵也保留，欄位檢視／篩選不受影響。
    """
    for alias, canonical in header_aliases().items():
        index = headers.get(alias)
        if index is not None:
            headers.setdefault(canonical.casefold(), index)


def is_blank_row(row: list) -> bool:
    """整列每一格 normalize 後都是空字串（NetSuite 匯出檔尾端常見的空白列）。"""
    return all(not normalize_text(value) for value in row)


def normalize_text(value: object) -> str:
    """VBA NormalizeText：IsError/IsNull/IsEmpty 視為空白、換行改空白、去頭尾空白。"""
    if value is None or is_excel_error_text(value):
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    text = str(value).replace("\r", " ").replace("\n", " ")
    return text.strip()


_DATE_PATTERNS = ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y")


def try_parse_date(value: object) -> date | None:
    """VBA TryParseDateValue 的移植：datetime、Excel 序號、文字日期、YYYYMMDD。"""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # 8 碼數字可能是 YYYYMMDD（例如某些欄位被存成數字型別而非文字），
        # 跟 Excel 序號範圍（最大 2958466，只有 7 碼）不重疊，先試著當
        # YYYYMMDD 解析；不是合法日期（例如月份超過 12）才退回序號判斷。
        if value == int(value) and 10000000 <= int(value) <= 99999999:
            int_value = int(value)
            try:
                return date(int_value // 10000, (int_value // 100) % 100, int_value % 100)
            except ValueError:
                pass
        if 0 < value < 2958466:
            return _from_serial(float(value))
        return None
    text = normalize_text(value)
    if not text:
        return None
    if "T" in text:
        text = text.split("T", 1)[0]
    if text.endswith(".0"):
        text = text[:-2]
    if text.isdigit() and len(text) != 8:
        serial = float(text)
        if 0 < serial < 2958466:
            return _from_serial(serial)
    if len(text) == 8 and text.isdigit():
        try:
            return date(int(text[:4]), int(text[4:6]), int(text[6:]))
        except ValueError:
            return None
    for pattern in _DATE_PATTERNS:
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    match = re.fullmatch(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})(?:\s.*)?", text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    return None


def _from_serial(serial: float) -> date | None:
    from datetime import timedelta

    base = datetime(1899, 12, 30)
    try:
        return (base + timedelta(days=serial)).date()
    except OverflowError:
        return None


def normalize_number(value: object) -> float:
    """VBA NormalizeNumber：數字直接轉，文字去千分位再轉，失敗回 0。"""
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0
