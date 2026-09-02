# -*- coding: utf-8 -*-
"""core 模組回歸測試。

執行方式（擇一）：
    python tests/test_core.py     # 不需安裝任何套件
    python -m pytest tests/       # 有裝 pytest 的話

涵蓋 2026-07 code review 修正的行為：批號解析失敗不合併、NetSuite 0 筆
結果、數字型別日期、Excel 錯誤值過濾、數量容差、到貨日關鍵字比對等。
"""
from __future__ import annotations

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from core import bundle_split
from core import inventory
from core import netsuite as netsuite_mod
from core import compare as compare_mod
from core import return_compare as return_compare_mod
from core.compare import _normalize_customer, _normalize_order_id
from core import export_log
from core import export_store
from core import gsheet
from core import table_filter
from core.inventory import _normalize_date, _normalize_item, _normalize_location, _qty_equal
from core.m00 import convert_rows as m00_convert_rows
from core.return_compare import _normalize_expiry as ret_normalize_expiry, _strip_dr_prefix
from core.shipping import (
    MEMO_SOURCE_LINE,
    describe_header_mapping,
    header_mapping_counts,
    MEMO_SOURCE_MAIN,
    MERGE_BY_PURCHASE_ORDER,
    MODE_ORDER,
    MODE_TRANSFER,
    _match_arrival_by_keyword,
    _normalize_arrival,
    _normalize_expiry as ship_normalize_expiry,
    _split_phone,
    _split_postal,
    convert_rows as shipping_convert_rows,
)
from core import xlio as xlio_mod
from core.xlio import (
    build_header_map,
    has_cjk,
    normalize_text,
    read_workbook,
    try_parse_date,
)

TEMPLATE_PATH = BASE_DIR / "mappings" / "HCT範本.xlsx"
M00_TEMPLATE_PATH = BASE_DIR / "mappings" / "M00出貨格式.xlsx"


# ------------------------------------------------------------------ xlio


def test_try_parse_date_variants():
    from datetime import date

    assert try_parse_date("2027-05-01") == date(2027, 5, 1)
    assert try_parse_date("2027/5/1") == date(2027, 5, 1)
    assert try_parse_date("20270501") == date(2027, 5, 1)
    # 數字型別的 8 碼日期（Excel 把日期欄存成 Number）
    assert try_parse_date(20270501.0) == date(2027, 5, 1)
    # Excel 序號
    assert try_parse_date(45000) == date(2023, 3, 15)
    assert try_parse_date("LOT-A1") is None
    assert try_parse_date("") is None


def test_normalize_text_excel_error_and_float():
    assert normalize_text("#REF!") == ""
    assert normalize_text(12345.0) == "12345"
    assert normalize_text(" a\nb ") == "a b"


def test_build_header_map_dedup_and_casefold():
    headers = build_header_map(["DR_料號", "dr_料號", "", None, "數量"])
    assert headers == {"dr_料號": 0, "數量": 4}


def test_read_workbook_utf8_bom_xml():
    xml = (
        '<?xml version="1.0"?>'
        '<Workbook xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">'
        '<ss:Worksheet ss:Name="S1"><ss:Table><ss:Row>'
        '<ss:Cell><ss:Data ss:Type="String">A</ss:Data></ss:Cell>'
        "</ss:Row></ss:Table></ss:Worksheet></Workbook>"
    ).encode("utf-8")
    book = read_workbook(b"\xef\xbb\xbf" + xml, "bom.xls")
    assert book["S1"][0][0] == "A"


# ------------------------------------------------------------------ 批號/效期 key


def test_shipping_expiry_unparseable_batches_stay_distinct():
    key_a, value_a, warning_a = ship_normalize_expiry("LOT-A1")
    key_b, _, _ = ship_normalize_expiry("LOT-B2")
    assert key_a != key_b
    assert key_a.startswith("raw:")
    assert value_a is None and warning_a
    assert ship_normalize_expiry("2027-05-01")[0] == "2027-05-01"
    assert ship_normalize_expiry("")[0] == ""


def test_return_expiry_same_text_matches_but_distinct_stays_apart():
    # 兩邊相同的無法解析文字要對得上（casefold）
    assert ret_normalize_expiry("待確認批號B")[0] == ret_normalize_expiry("待確認批號b")[0]
    # 不同文字不能落到同一 key
    assert ret_normalize_expiry("待確認批號B")[0] != ret_normalize_expiry("效期未確認")[0]
    # 可解析日期輸出 (key, 效期 YYYY-MM-DD, 批號 yyyymmdd)
    assert ret_normalize_expiry("20270501") == ("2027-05-01", "2027-05-01", "20270501")
    assert ret_normalize_expiry("2027/5/1") == ("2027-05-01", "2027-05-01", "20270501")
    # 無法解析的批號:效期與批號都照抄原文
    assert ret_normalize_expiry("LOT-A")[1:] == ("LOT-A", "LOT-A")
    assert ret_normalize_expiry(None) == ("", "", "")


# ------------------------------------------------------------------ inventory


def test_inventory_normalize_date_numeric_and_text():
    assert _normalize_date(20270101.0) == ("20270101", True)
    assert _normalize_date("2027/1/1") == ("20270101", True)
    assert _normalize_date(None) == ("", True)
    assert _normalize_date("   ") == ("", True)
    assert _normalize_date("abc") == ("", False)
    assert _normalize_date(45000.0)[1] is True  # Excel 序號


def test_inventory_excel_error_treated_as_blank():
    assert _normalize_item("#REF!") == ""
    assert _normalize_location("#N/A") == ""
    assert _normalize_location("g10_x") == "G10"


def test_inventory_quantity_tolerance():
    assert _qty_equal(0.1 + 0.2, 0.3)
    assert not _qty_equal(10.0, 10.5)


def test_inventory_contract_reconcile_against_ns293():
    """代工廠庫存核對報表 × NetSuite:排除合計列、單一數量欄當兩用。

    倉別不再限定:E01 與 NetSuite 側的 G10 都照常核對，只有「合計／總計」
    小計列(倉別與品號皆空白)仍列入排除數。
    """
    contract = _spreadsheetml([
        ["庫存日期", "倉別", "倉別名稱", "品號", "品名", "批號", "庫存數量"],
        ["20260724", "D01", "凱芬妮倉", "300020400025", "品A", "", "100"],
        ["20260724", "D01", "凱芬妮倉", "300020400026", "品B", "", "50"],
        ["", "", "凱芬妮倉 合計", "", "", "", "150"],
        ["20260724", "E01", "非代工廠倉", "300020400027", "品C", "", "10"],
        ["", "", "總計", "", "", "", "160"],
    ])
    ns293 = _spreadsheetml([
        ["地點", "到期日", "項目", "庫存數 總和"],
        ["D01_凱芬妮", "", "300020400025_X", "90"],
        ["D01_凱芬妮", "", "300020400026_X", "50"],
        ["G10", "", "300020400028_X", "5"],
    ])
    assert inventory.detect_source(contract, "c.xls") == inventory.SYSTEM_CONTRACT

    result = inventory.reconcile(ns293, "ns.xls", contract, "c.xls")
    assert result.ext_label == inventory.SYSTEM_CONTRACT
    assert result.output_name.startswith("代工廠庫存核對結果_")
    # 代工廠側只剩合計/總計列 2 筆;NetSuite 側不再排除 G10
    assert result.hct_stats.excluded_rows == 2
    assert result.ns_stats.excluded_rows == 0
    assert result.anomalies == []
    by_item = {r["料號"]: r for r in result.item_rows}
    assert by_item["300020400025"]["代工廠 庫存數量"] == 100
    assert by_item["300020400025"]["總庫存差額"] == 10
    assert by_item["300020400025"]["狀態"] == inventory.STATUS_BOTH_DIFF
    assert by_item["300020400026"]["狀態"] == inventory.STATUS_MATCH
    # 先前被倉別過濾掉的兩筆現在都會出現
    assert by_item["300020400027"]["狀態"] == "僅 代工廠 存在"   # E01 倉
    assert by_item["300020400028"]["狀態"] == inventory.STATUS_NS_ONLY  # G10 倉
    assert "僅 代工廠 存在" in result.statuses


def test_inventory_ns_lot_format_reconcile_against_hct():
    """NetSuite 批號版 × HCT:倉別代碼直接當倉別、批號當效期、在庫數量兩用。

    核對鍵是 倉別＋料號＋效期;倉別不限定，G15 這種倉別照常核對。
    """
    ns_lot = _spreadsheetml([
        ["料號", "品名", "倉別代碼", "倉別名稱", "批號", "在庫數量"],
        ["101021190002", "潔顏凝露50ML", "G10", "新竹-正常良品倉", "20290401", "7980"],
        ["101021190002", "潔顏凝露50ML", "G10", "新竹-正常良品倉", "20290701", "20"],
        ["101011280002", "卸妝凝露200ML", "G80", "新竹-待報廢倉", "20260701", "1"],
        ["109000000001", "低效品", "G15", "新竹-低效良品倉", "20270301", "9"],
    ])
    assert inventory.detect_source(ns_lot, "ns.xlsx") == inventory.SYSTEM_NETSUITE_LOT

    hct = _spreadsheetml([
        ["儲區類別", "客戶產品編號", "產品名稱", "有效日期", "可出數量", "庫存數量"],
        ["G10", "101021190002", "潔顏凝露50ML", "20290401", "7975", "7975"],
        ["G10", "101021190002", "潔顏凝露50ML", "20290701", "20", "20"],
        ["G80", "101011280002", "卸妝凝露200ML", "20260701", "1", "1"],
    ])

    result = inventory.reconcile(ns_lot, "ns.xlsx", hct, "h.xls")
    assert result.ext_label == inventory.SYSTEM_HCT
    assert result.ns_stats.valid_rows == 4
    assert result.ns_stats.excluded_rows == 0   # 不再因倉別排除 G15
    assert result.anomalies == []
    by_location = {r["倉別"]: r for r in result.item_rows}
    assert by_location["G15"]["狀態"] == inventory.STATUS_NS_ONLY

    # 明細鍵 = 倉別＋料號＋效期:同料號同倉別的兩個批號要分開兩列
    lots = [r for r in result.detail_rows if r["料號"] == "101021190002"]
    assert len(lots) == 2
    by_expiry = {str(r["到期日"]): r for r in lots}
    assert by_expiry["2029-04-01"]["NetSuite 數量"] == 7980
    assert by_expiry["2029-04-01"]["總庫存差額"] == -5
    assert by_expiry["2029-04-01"]["狀態"] == inventory.STATUS_BOTH_DIFF
    assert by_expiry["2029-07-01"]["狀態"] == inventory.STATUS_MATCH

    # 料號彙總層則把兩個批號加總成一列
    by_item = {r["料號"]: r for r in result.item_rows}
    assert by_item["101021190002"]["NetSuite 數量"] == 8000
    # 在庫數量單欄兩用:可用量與總庫存量取同一個值
    assert by_item["101011280002"]["NetSuite 項目計數"] == 1
    assert by_item["101011280002"]["狀態"] == inventory.STATUS_MATCH


def test_inventory_ns_lot_does_not_collide_with_contract_format():
    """代工廠報表也有 倉別名稱／品名／批號,不可被誤判成 NetSuite 批號版。"""
    contract = _spreadsheetml([
        ["庫存日期", "倉別", "倉別名稱", "品號", "品名", "批號", "庫存數量"],
        ["20260724", "D01", "凱芬妮倉", "300020400025", "品A", "", "100"],
    ])
    assert inventory.detect_source(contract, "c.xls") == inventory.SYSTEM_CONTRACT


def test_inventory_m00_reconcile_sums_qty_plus_reserved():
    """M00 庫存詳情 × NetSuite：庫存量＝數量＋組合保留，整份視為 M00 倉。"""
    m00 = _spreadsheetml([
        ["SKU", "國際條碼", "商品名稱", "有效日期", "商品狀態", "狀態", "數量", "組合保留", "批號"],
        ["115101040013", "471", "面膜4入", "2029-01-01", "良好", "可以出倉", "377", "0", "L1"],
        ["115101040013", "471", "面膜4入", "2029-01-01", "良好", "準備出倉", "25", "4", "L1"],
        ["101051190015", "471", "精華乳", "2027-06-01", "NG", "可以出倉", "10", "2", "L2"],
    ])
    ns293 = _spreadsheetml([
        ["地點", "到期日", "項目", "庫存數 總和"],
        ["M00_電商倉", "2029/1/1", "115101040013_X", "406"],
        ["G10", "", "999000000001_X", "5"],
    ])
    assert inventory.detect_source(m00, "m00.xlsx") == inventory.SYSTEM_M00

    result = inventory.reconcile(m00, "m00.xlsx", ns293, "ns.xls")
    assert result.ext_label == inventory.SYSTEM_M00
    assert result.output_name.startswith("M00庫存核對結果_")
    assert result.ns_stats.excluded_rows == 1   # 非 M00 倉（G10）
    assert result.hct_stats.excluded_rows == 0
    by_item = {r["料號"]: r for r in result.item_rows}
    assert by_item["115101040013"]["M00 庫存數量"] == 406   # 377+0+25+4
    assert by_item["115101040013"]["狀態"] == inventory.STATUS_MATCH
    assert by_item["101051190015"]["M00 庫存數量"] == 12    # 10+2
    assert by_item["101051190015"]["狀態"] == "僅 M00 存在"
    assert "僅 M00 存在" in result.statuses


def test_inventory_m00_location_variants_and_dash_expiry():
    """NetSuite 倉別 M00倉/M001（無底線）要收斂成 M00 鍵；全形破折號＝無到期日。"""
    m00 = _spreadsheetml([
        ["SKU", "商品名稱", "有效日期", "數量", "組合保留", "批號"],
        ["115101040013", "面膜4入", "2029-01-01", "100", "0", "L1"],
        ["888000000001", "贈品帆布袋", "—", "30", "5", ""],  # 全形破折號＝無到期日
    ])
    ns293 = _spreadsheetml([
        ["地點", "到期日", "項目", "庫存數 總和"],
        ["M00倉", "2029/1/1", "115101040013_X", "100"],   # 無底線的 M00 開頭倉別
        ["M001", "", "888000000001_X", "35"],
    ])
    result = inventory.reconcile(m00, "m00.xlsx", ns293, "ns.xls")
    assert result.hct_stats.anomaly_rows == 0  # 破折號不能被當成日期異常排除
    assert result.ns_stats.excluded_rows == 0
    by_item = {r["料號"]: r for r in result.item_rows}
    assert len(by_item) == 2  # 兩邊鍵收斂成 M00，不能拆成僅單邊存在的假差異
    assert by_item["115101040013"]["狀態"] == inventory.STATUS_MATCH
    assert by_item["888000000001"]["M00 庫存數量"] == 35  # 30+5
    assert by_item["888000000001"]["狀態"] == inventory.STATUS_MATCH
    assert all(r["倉別"] == "M00" for r in result.item_rows)


def test_inventory_hct_reconcile_output_unchanged():
    """原本的 HCT × NetSuite 流程不受新格式影響(欄名、檔名、狀態清單)。

    倉別不再限定:Z99 這種非 G 倉不再被排除，而是照常核對成「僅 HCT 存在」。
    """
    hct = _spreadsheetml([
        ["儲區類別", "客戶產品編號", "有效日期", "可出數量", "庫存數量", "產品名稱"],
        ["G10", "SKU1", "20270501", "10", "12", "品A"],
        ["Z99", "SKU2", "", "1", "1", "非G倉也要核對"],
    ])
    ns = _spreadsheetml([
        ["地點", "到期日", "DR_料號", "項目計數 總和", "數量 總和"],
        ["G10_X", "2027/5/1", "SKU1", "10", "12"],
    ])
    result = inventory.reconcile(hct, "hct.xls", ns, "ns.xls")
    assert result.ext_label == inventory.SYSTEM_HCT
    assert result.output_name.startswith("庫存核對結果_")
    assert result.statuses == inventory.ALL_STATUSES
    assert result.hct_stats.excluded_rows == 0      # 不再因倉別排除任何列
    assert result.hct_stats.valid_rows == 2
    by_item = {r["料號"]: r for r in result.item_rows}
    assert by_item["SKU1"]["HCT 可出數量"] == 10
    assert by_item["SKU1"]["狀態"] == inventory.STATUS_MATCH
    assert by_item["SKU2"]["倉別"] == "Z99"
    assert by_item["SKU2"]["狀態"] == "僅 HCT 存在"


# ------------------------------------------------------------------ 到貨日欄位比對


def test_arrival_keyword_excludes_actual_and_warns():
    headers = {"實際到貨日": 0, "客戶指定到貨日": 1}
    notes = _match_arrival_by_keyword(headers)
    assert headers["dr_預計到貨日期"] == 1
    assert notes and "客戶指定到貨日" in notes[0]


def test_arrival_blank_warning_mentions_merge_only_in_order_mode():
    # 訂單模式（真的有跨單合併）才提「不跨單合併」；M00/調撥單用中性訊息
    _, order_warning = _normalize_arrival(None, merge_hint=True)
    _, neutral_warning = _normalize_arrival(None)
    assert "不跨單合併" in order_warning
    assert "不跨單合併" not in neutral_warning
    assert "空白" in neutral_warning


def test_arrival_keyword_no_match_when_only_excluded():
    headers = {"實際到貨日": 0}
    notes = _match_arrival_by_keyword(headers)
    assert "dr_預計到貨日期" not in headers
    assert notes == []


# ------------------------------------------------------------------ compare 正規化


def test_netsuite_saved_search_english_headers_map_to_converter_names():
    """saved search 沒設中文 Label 時回傳英文預設名稱,要譯成轉換認得的欄名。

    欄名取自正式區實際回傳(調撥單未出貨明細 searchId=159):品名欄是
    Display Name,過去不在對照表裡,轉出的商品名稱就整欄空白;批號與主要
    備忘錄同樣是英文欄名。銷售訂單那支(customsearch517)欄名已是中文,
    不受影響。
    """
    transfer_columns = [
        "Internal ID", "Document Number", "Date", "DR_料號", "Display Name",
        "Transaction Serial/Lot Number", "來源倉別", "目標倉別", "出貨數量",
        "倉儲", "倉儲聯繫人", "倉儲電話", "倉儲地址", "Memo (Main)", "Item",
        "來源地點", "目標地點", "DR_預計到貨日期", "DR_預計到貨時段",
    ]
    mapped = netsuite_mod._normalize_headers([transfer_columns])[0]
    for want in ("內部 ID", "文件編號", "日期", "項目", "顯示名稱",
                 "交易序號/批號", "備忘錄 (主要)"):
        assert want in mapped, want
    # shipping 的 MODE_TRANSFER 別名把「顯示名稱」接到「項目名稱」,品名才有值
    assert "顯示名稱" in mapped and "Display Name" not in mapped

    # 已是中文欄名的 saved search 不能被動到
    order_columns = ["內部 ID", "文件編號", "日期", "DR_料號", "項目名稱",
                     "交易序號/批號", "出貨客戶", "備忘錄", "項目"]
    assert netsuite_mod._normalize_headers([order_columns])[0] == order_columns


def test_netsuite_client_rejects_blank_credentials():
    """金鑰空白時要當場講清楚。

    空金鑰簽出來的 Authorization header 是 oauth_consumer_key=""，NetSuite 回
    400 INVALID_REQUEST——跟「script id 不存在」的錯誤訊息一字不差，不擋的話
    會被誤判成 RESTlet 部署壞掉。
    """
    cfg = {
        "account_id": "11776409", "consumer_key": "", "consumer_secret": "  ",
        "token_id": "tk", "token_secret": "ts",
    }
    try:
        netsuite_mod.NetSuiteClient(cfg)
    except netsuite_mod.NetSuiteError as exc:
        assert "consumer_key" in str(exc) and "consumer_secret" in str(exc)
    else:
        raise AssertionError("金鑰空白卻沒有擋下來")
    cfg.update(consumer_key="ck", consumer_secret="cs")
    assert netsuite_mod.NetSuiteClient(cfg).account == "11776409"


def test_learn_header_aliases_joins_chinese_and_english_labels():
    """自動對照:同一個欄位內部 ID 在 A 搜尋有中文 Label、在 B 搜尋是英文預設名稱,
    兩邊 join 起來就知道英文欄名該當成哪個中文欄名,不用人工維護對照表。
    """
    chinese = [
        {"name": "tranid", "label": "文件編號"},
        {"name": "memomain", "label": "備忘錄 (主要)"},
        {"name": "serialnumbers", "label": "交易序號/批號"},
        {"name": "formulatext", "label": "來源倉別"},
        {"name": "quantity", "label": "出貨數量"},
    ]
    english = [
        {"name": "tranid", "label": "Document Number"},
        {"name": "memomain", "label": "Memo (Main)"},
        {"name": "serialnumbers", "label": "Transaction Serial/Lot Number"},
        {"name": "formulatext", "label": "Source Warehouse"},
        {"name": "quantity", "label": None},          # 連 Label 都沒設
    ]
    aliases, rows = netsuite_mod.learn_header_aliases(
        [("中文那支", chinese), ("英文那支", english)]
    )
    assert aliases["document number"] == "文件編號"
    assert aliases["memo (main)"] == "備忘錄 (主要)"
    assert aliases["transaction serial/lot number"] == "交易序號/批號"
    # 欄位內部 ID 本身也要學起來(Label 完全沒設時 RESTlet 回的就是它)
    assert aliases["tranid"] == "文件編號"
    assert aliases["quantity"] == "出貨數量"
    # 公式欄的內部 ID 不同欄會撞在一起,刻意不學
    assert "source warehouse" not in aliases and "formulatext" not in aliases
    # 對照鍵永遠不含中文:來源檔本身的中文欄名不會被自動對照改掉
    assert not any(has_cjk(key) for key in aliases)
    assert {r["對應中文欄名"] for r in rows} == {"文件編號", "備忘錄 (主要)", "交易序號/批號", "出貨數量"}


def test_learn_header_aliases_prefers_label_the_converter_knows():
    """同一個內部 ID 被不同搜尋取了不同中文名時,優先採用工具本來就認得的那個。"""
    aliases, _rows = netsuite_mod.learn_header_aliases([
        ("自己取名那支", [{"name": "tranid", "label": "單據號碼"}]),
        ("標準那支", [{"name": "tranid", "label": "文件編號"}]),
        ("英文那支", [{"name": "tranid", "label": "Document Number"}]),
    ])
    assert aliases["document number"] == "文件編號"


def test_restlet_column_metadata_label_wins_but_internal_id_rescues():
    """新版 RESTlet 回 {name,label} 時的欄名決定順序:

    1. 有 Label 就以 Label 為準——自訂中文標籤(出貨數量)不可以被內部 ID 的
       通用譯名(quantity -> 數量)蓋掉,否則轉換反而找不到欄位。
    2. Label 是英文預設名稱時照對照表譯成中文。
    3. Label 認不得、但內部 ID 對得上某個必要欄位且整份報表都沒有該欄時,
       才用內部 ID 補救。
    4. 舊版 RESTlet 只回字串,行為要跟以前完全一樣。
    """
    header = netsuite_mod._normalize_headers([[
        {"name": "quantity", "label": "出貨數量"},
        {"name": "tranid", "label": "Document Number"},
        {"name": "trandate", "label": "Posting Period Date"},
        {"name": "displayname", "label": None},
    ]])[0]
    assert header[0] == "出貨數量"      # 自訂中文標籤原樣保留
    assert header[1] == "文件編號"      # 英文預設名稱譯成中文
    assert header[2] == "日期"          # 認不得的 Label + 內部 ID 對得上必要欄位
    assert header[3] == "顯示名稱"      # 沒有 Label 時退到內部 ID

    # 內部 ID 補救只在該必要欄位真的缺席時才介入
    header = netsuite_mod._normalize_headers([[
        {"name": "trandate", "label": "Posting Period Date"},
        {"name": "custcol_x", "label": "日期"},
    ]])[0]
    assert header == ["Posting Period Date", "日期"]

    # 舊版 RESTlet(字串陣列)
    assert netsuite_mod._normalize_headers([["Document Number", "出貨數量"]])[0] == [
        "文件編號", "出貨數量",
    ]


def test_header_alias_cache_extends_upload_path():
    """學回來的對照存成檔案,是為了讓沒連 NetSuite 的「上傳檔案」路徑也吃得到。"""
    import tempfile

    original = xlio_mod.HEADER_ALIAS_CACHE_PATH
    with tempfile.TemporaryDirectory() as tmp:
        xlio_mod.HEADER_ALIAS_CACHE_PATH = Path(tmp) / "cache.json"
        try:
            xlio_mod.save_header_alias_cache({"Lot Number": "交易序號/批號"}, [])
            headers = build_header_map(["Lot Number", "出貨數量"])
            xlio_mod.apply_netsuite_header_aliases(headers)
            assert headers["交易序號/批號".casefold()] == 0
            # 內建表優先:同一個鍵兩邊都有時以人工驗證過的為準
            xlio_mod.save_header_alias_cache({"date": "亂七八糟"}, [])
            assert xlio_mod.header_aliases()["date"] == "日期"
        finally:
            xlio_mod.HEADER_ALIAS_CACHE_PATH = original
            xlio_mod.load_header_alias_cache(refresh=True)


def test_describe_header_mapping_flags_silent_gaps():
    """轉換前的欄位對照預檢:自動對應要看得出來,選用欄缺少要講後果(不是靜默空白)。"""
    header = [
        "內部 ID", "文件編號", "日期", "DR_料號", "顯示名稱",
        "Transaction Serial/Lot Number", "來源倉別", "出貨數量", "倉儲",
        "倉儲聯繫人", "倉儲電話", "倉儲地址", "Memo (Main)", "Item",
        "目標地點", "DR_預計到貨日期", "DR_預計到貨時段",
    ]
    rows = {r["欄位"]: r for r in describe_header_mapping(
        header, MODE_TRANSFER, MEMO_SOURCE_MAIN, MERGE_BY_PURCHASE_ORDER
    )}
    assert rows["序號/批號"]["來源欄位"] == "Transaction Serial/Lot Number"
    assert rows["序號/批號"]["狀態"] == "✅ 已自動對應"
    assert rows["DR_溫馨提醒"]["來源欄位"] == "Memo (Main)"
    assert rows["文件編號"]["狀態"] == "✅ 對應"
    assert rows["虛擬倉來源"]["來源欄位"] == "來源倉別"
    # 沒有的選用欄要明講後果,不能只說「缺少」
    assert rows["客戶採購單編號"]["狀態"].startswith("⚠️ 缺少 — 無法以採購單合併")
    missing_required, missing_optional = header_mapping_counts(list(rows.values()))
    assert missing_required == 0 and missing_optional >= 1

    # 必要欄真的缺就要標成無法轉換
    broken = [h for h in header if h not in ("文件編號", "日期")]
    rows = {r["欄位"]: r for r in describe_header_mapping(broken, MODE_TRANSFER)}
    assert rows["日期"]["狀態"].startswith("❌")
    assert header_mapping_counts(list(rows.values()))[0] == 2


def test_upload_transfer_report_with_english_default_headers():
    """手動匯出的調撥單報表中英混雜時,上傳路徑也要認得英文預設欄名。

    欄名取自實際匯出檔(調撥單未出貨明細物流專用):有設中文 Label 的欄
    (內部 ID/文件編號/日期/來源倉別…)是中文,沒設的退回英文
    (Transaction Serial/Lot Number、Memo (Main)、Transaction Line Type)。
    過去英中對照只掛在 RESTlet 直接抓取那條路徑,所以同一支 saved search
    「直接抓取」會過、「下載成檔案再上傳」卻出事,而且兩種出事方式都要擋:
      1. 文件編號/日期/項目 是英文 -> 必要欄檢查直接擋下(來源檔缺少必要欄位)
      2. 批號/備忘錄 是英文 -> 不擋,默默輸出空白效期與空白備註
    """
    header = [
        "Internal ID", "Document Number", "Date", "DR_料號", "顯示名稱",
        "Transaction Serial/Lot Number", "來源倉別", "目標倉別", "出貨數量",
        "倉儲", "倉儲聯繫人", "倉儲電話", "倉儲地址", "Memo (Main)",
        "Transaction Line Type", "Item", "來源地點", "目標地點",
        "DR_預計到貨日期", "DR_預計到貨時段",
    ]
    row = [
        "75939", "TO-000264", "2026-09-02", "101031250014", "玻尿酸保濕精華化妝水150ML",
        "20290801", "G30 台北倉", "C13", "24",
        "C13", "Ariel", "0225179888", "104 台北市中山區南京東路三段70號13樓", "0901進貨QB",
        "出貨", "101031250014_HHT150A", "G30_台北倉", "C13_達爾膚-品檢倉",
        "2026-09-03", "早(9-12)",
    ]
    result = shipping_convert_rows(
        [header, row], MODE_TRANSFER, TEMPLATE_PATH, memo_source=MEMO_SOURCE_MAIN,
    )
    assert result.output_items == 1
    assert _output_cell(result, 2, 1) == "TO-000264"          # Document Number
    assert _output_cell(result, 2, 3) == "101031250014"        # Item(料號空白時的備援來源)
    assert _output_cell(result, 2, 7) == "20290801"            # 批號 <- Transaction Serial/Lot Number
    assert str(_output_cell(result, 2, 33)).startswith("2029-08-01")   # 效期
    assert "0901進貨QB" in _output_cell(result, 2, 19)          # 備註 <- Memo (Main)
    assert _output_cell(result, 2, 30) == "G30"                # 來源倉別仍正常


def test_compare_quantity_key_uses_normalized_expiry():
    """數量核對鍵＝客戶+料號+效期:兩邊效期寫法不同要視為同一組,不是兩筆假差異。

    兩份來源都沒有倉別欄,所以鍵停在客戶層級(同料號同效期出給不同客戶
    的量不能合併)。
    """
    unshipped = _spreadsheetml([
        ["出貨客戶", "DR_料號", "交易序號/批號", "出貨數量", "客戶採購單編號"],
        ["屈臣氏", "101021190002", "20270501", "10", "PO-1"],
        ["寶雅", "101021190002", "20270501", "4", "PO-2"],
    ])
    delivery = _spreadsheetml([
        ["客戶簡稱", "品號", "批號", "銷貨數量", "網路訂單編號"],
        ["屈臣氏", "101021190002", "2027/5/1", "10", "PO-1"],
        ["寶雅", "101021190002", "2027-05-01", "4", "PO-2"],
    ])
    result = compare_mod.compare(unshipped, "u.xls", delivery, "d.xls")
    # 效期正規化前這裡會是 4 筆(每邊各自成組);正規化後兩邊對得上剩 2 筆
    assert result.quantity_rows == 2
    assert result.quantity_status_counts == {"一致": 2}


def test_compare_unparseable_batches_stay_distinct():
    """無法解析成日期的批號保留原文當鍵,不同批號不可被併成同一筆。"""
    assert compare_mod._normalize_expiry("LOT-A")[0] != compare_mod._normalize_expiry("LOT-B")[0]
    assert compare_mod._normalize_expiry("LOT-A")[0] == compare_mod._normalize_expiry("lot-a")[0]
    # 空效期用空字串鍵:代表這列不在乎效期,兩邊的空效期列仍互相比對
    assert compare_mod._normalize_expiry("") == ("", "", "")
    # (key, 效期 YYYY-MM-DD, 批號 yyyymmdd)
    assert compare_mod._normalize_expiry("20270501") == ("2027-05-01", "2027-05-01", "20270501")
    assert compare_mod._normalize_expiry("2027/5/1") == ("2027-05-01", "2027-05-01", "20270501")
    assert compare_mod._normalize_expiry("LOT-A")[1:] == ("LOT-A", "LOT-A")


def test_return_compare_key_includes_location():
    """退貨核對鍵＝料號+效期+倉別:同料號同效期不同倉別要分開兩列。

    RA 側倉別是「G10_新竹…」帶底線的地點字串,HCT 入庫是純代碼「G10」,
    要取底線前段才對得上。
    """
    ra = _spreadsheetml([
        ["文件編號", "客戶", "DR_料號", "交易序號/批號", "交易序號/批號數量", "倉別", "顯示名稱"],
        ["RA-DW-1", "屈臣氏", "101021190002", "20270501", "10", "G10_新竹正常良品倉", "潔顏凝露"],
        ["RA-DW-2", "屈臣氏", "101021190002", "20270501", "3", "G30_新竹瑕疵倉", "潔顏凝露"],
    ])
    hct = _spreadsheetml([
        ["退貨單號", "產品編號", "效期", "數量", "儲區類別", "產品名稱"],
        ["R1", "DR101021190002", "2027/5/1", "10", "G10", "潔顏凝露"],
    ])
    result = return_compare_mod.compare(ra, "ra.xls", hct, "h.xls")
    # G10 兩邊對上、G30 只有退貨授權 → 2 筆而不是併成 1 筆
    assert result.total_rows == 2
    assert result.status_counts == {"一致": 1, "僅退貨授權": 1}


def test_compare_result_keeps_batch_column():
    """數量核對結果要留批號欄:效期給 yyyy-mm-dd,批號給 yyyymmdd 原始寫法。"""
    unshipped = _spreadsheetml([
        ["出貨客戶", "DR_料號", "交易序號/批號", "出貨數量", "客戶採購單編號"],
        ["屈臣氏", "101021190002", "20270501", "10", "PO-1"],
        ["屈臣氏", "101021190003", "LOT-A", "2", "PO-1"],
    ])
    delivery = _spreadsheetml([
        ["客戶簡稱", "品號", "批號", "銷貨數量", "網路訂單編號"],
        ["屈臣氏", "101021190002", "2027/5/1", "10", "PO-1"],
    ])
    result = compare_mod.compare(unshipped, "u.xls", delivery, "d.xls")
    assert "批號" in compare_mod._QUANTITY_COLUMNS
    rows = _read_sheet_rows(result.output_bytes, "數量核對")
    header = rows[0]
    batch_col = header.index("批號")
    by_item = {r[header.index("料號")]: r for r in rows[1:]}
    assert by_item["101021190002"][header.index("效期")] == "2027-05-01"
    assert by_item["101021190002"][batch_col] == "20270501"   # 文字,不是 20270501 數字
    assert by_item["101021190003"][batch_col] == "LOT-A"


def test_return_compare_result_keeps_batch_column():
    ra = _spreadsheetml([
        ["文件編號", "客戶", "DR_料號", "交易序號/批號", "交易序號/批號數量", "倉別", "顯示名稱"],
        ["RA-DW-1", "屈臣氏", "101021190002", "20270501", "10", "G10_新竹正常良品倉", "潔顏凝露"],
    ])
    hct = _spreadsheetml([
        ["退貨單號", "產品編號", "效期", "數量", "儲區類別", "產品名稱"],
        ["R1", "DR101021190002", "2027/5/1", "10", "G10", "潔顏凝露"],
    ])
    result = return_compare_mod.compare(ra, "ra.xls", hct, "h.xls")
    rows = _read_sheet_rows(result.output_bytes, "退貨核對明細")
    header = rows[0]
    assert rows[1][header.index("效期")] == "2027-05-01"
    assert rows[1][header.index("批號")] == "20270501"


def test_inventory_detail_keeps_batch_column():
    """日期明細留批號欄(yyyymmdd 文字);料號彙總沒有到期日,也不該有批號欄。"""
    ns_lot = _spreadsheetml([
        ["料號", "品名", "倉別代碼", "倉別名稱", "批號", "在庫數量"],
        ["101021190002", "潔顏凝露50ML", "G10", "新竹-正常良品倉", "20290401", "20"],
        ["109000000002", "無到期日贈品", "G10", "新竹-正常良品倉", "", "5"],
    ])
    hct = _spreadsheetml([
        ["儲區類別", "客戶產品編號", "產品名稱", "有效日期", "可出數量", "庫存數量"],
        ["G10", "101021190002", "潔顏凝露50ML", "2029/4/1", "20", "20"],
        ["G10", "109000000002", "無到期日贈品", "", "5", "5"],
    ])
    result = inventory.reconcile(ns_lot, "ns.xlsx", hct, "h.xls")
    by_item = {r["料號"]: r for r in result.detail_rows}
    assert by_item["101021190002"]["批號"] == "20290401"
    assert by_item["109000000002"]["批號"] == "無到期日"
    assert "批號" not in result.item_rows[0]

    rows = _read_sheet_rows(result.output_bytes, "日期明細")
    header = rows[0]
    assert header.index("批號") == header.index("到期日") + 1
    detail = {r[header.index("料號")]: r for r in rows[1:]}
    assert detail["101021190002"][header.index("批號")] == "20290401"
    item_header = _read_sheet_rows(result.output_bytes, "料號彙總")[0]
    assert "批號" not in item_header
    # 狀態欄位置隨新欄往後移,底色仍要標在狀態欄上
    assert header.index("狀態") == len(header) - 4


def test_return_compare_location_normalizer():
    assert return_compare_mod._normalize_location("g10_新竹正常良品倉") == "G10"
    assert return_compare_mod._normalize_location(" G30 ") == "G30"
    assert return_compare_mod._normalize_location(None) == ""


def test_compare_normalizers():
    assert _normalize_customer("CUS-001 康是美") == "康是美"
    assert _normalize_order_id("SO12345-1") == "SO12345"
    assert _normalize_order_id("-abc") == ""
    assert _strip_dr_prefix("DR902224") == "902224"
    assert _strip_dr_prefix("DRW-X") == "DRW-X"


# ------------------------------------------------------------------ 電話/地址


def test_split_phone_and_postal():
    assert _split_phone("+886-912-345-678") == ("", "0912345678")
    assert _split_phone("02-2345-6789") == ("0223456789", "")
    assert _split_postal("10041台北市中正區路1號") == ("10041", "台北市中正區路1號")
    assert _split_postal("台北市") == ("", "台北市")


# ------------------------------------------------------------------ 端對端轉換


_E2E_HEADER = [
    "內部 ID", "文件編號", "銷售訂單類型", "日期", "項目", "項目名稱", "DR_料號",
    "序號/批號", "出貨客戶", "出貨數量", "門市/倉儲", "門市/倉儲聯繫人",
    "門市/倉儲電話", "門市/倉儲地址", "DR_預計到貨日期", "DR_預計到貨時段",
]


def _e2e_row(batch: str, qty: float, slot: str = "早(9-12)") -> list:
    return [
        "1", "SO-1", "一般銷售訂單", "2026-07-01", "SKU001_X", "品名A", "SKU001",
        batch, "客戶A", qty, "門市A", "聯絡人", "0912345678",
        "10041台北市中正區路1號", "2026-07-30", slot,
    ]


def test_e2e_distinct_bad_batches_not_merged():
    rows = [_E2E_HEADER, _e2e_row("LOT-A1", 10), _e2e_row("LOT-B2", 5)]
    result = shipping_convert_rows(rows, MODE_ORDER, TEMPLATE_PATH)
    assert result.output_items == 2
    m00_result = m00_convert_rows(rows, MODE_ORDER, M00_TEMPLATE_PATH)
    assert m00_result.output_items == 2


def test_e2e_same_batch_still_merges():
    rows = [_E2E_HEADER, _e2e_row("2027-05-01", 10), _e2e_row("2027-05-01", 5)]
    result = shipping_convert_rows(rows, MODE_ORDER, TEMPLATE_PATH)
    assert result.output_items == 1


def test_e2e_unknown_time_slot_warns():
    rows = [_E2E_HEADER, _e2e_row("2027-05-01", 10, slot="晚（17-20）")]
    m00_result = m00_convert_rows(rows, MODE_ORDER, M00_TEMPLATE_PATH)
    assert any("無法對應" in w for w in m00_result.warnings)


# ------------------------------------------------------------------ 採購單合併


def _po_row(doc: str, order_type: str, material: str = "SKU001",
            address: str = "10041台北市中正區路1號") -> list:
    return [
        "1", doc, order_type, "2026-07-01", f"{material}_X", "品名A", material,
        "2027-05-01", "客戶A", 10, "門市A", "聯絡人", "0912345678",
        address, "2026-07-30", "早(9-12)",
    ]


_PO_HEADER = _E2E_HEADER + ["客戶採購單編號"]


def test_merge_by_po_merges_regardless_of_order_type():
    """同採購單編號的訂單合併成一張送貨單，不看訂單類型（全備品出貨也合併）。"""
    rows = [
        _PO_HEADER,
        _po_row("SO-1", "備品出貨") + ["PO-100"],
        _po_row("SO-2", "備品出貨", material="SKU002") + ["PO-100"],
        _po_row("SO-3", "備品出貨") + ["PO-200"],
    ]
    result = shipping_convert_rows(
        rows, MODE_ORDER, TEMPLATE_PATH, merge_by=MERGE_BY_PURCHASE_ORDER
    )
    assert result.shipments == 2
    assert result.merged_groups == 1
    # 合併單的送貨單號/訂單編號列出兩張 SO（沒有一般銷售訂單也不能空白）
    assert _output_cell(result, 2, 1) == "SO-1、SO-2"
    assert _output_cell(result, 2, 2) == "SO-1、SO-2"
    # 原規則下同樣資料完全不合併（無一般銷售訂單）
    baseline = shipping_convert_rows(rows, MODE_ORDER, TEMPLATE_PATH)
    assert baseline.shipments == 3
    assert baseline.merged_groups == 0


def test_merge_by_po_blank_po_stays_separate_and_warns():
    rows = [
        _PO_HEADER,
        _po_row("SO-1", "備品出貨") + [""],
        _po_row("SO-2", "備品出貨", material="SKU002") + [""],
    ]
    result = shipping_convert_rows(
        rows, MODE_ORDER, TEMPLATE_PATH, merge_by=MERGE_BY_PURCHASE_ORDER
    )
    assert result.shipments == 2
    assert any("客戶採購單編號空白" in w for w in result.warnings)


def test_merge_by_po_warns_on_inconsistent_recipient():
    rows = [
        _PO_HEADER,
        _po_row("SO-1", "備品出貨") + ["PO-100"],
        _po_row("SO-2", "備品出貨", material="SKU002",
                address="20041基隆市另一個地址2號") + ["PO-100"],
    ]
    result = shipping_convert_rows(
        rows, MODE_ORDER, TEMPLATE_PATH, merge_by=MERGE_BY_PURCHASE_ORDER
    )
    assert result.shipments == 1
    assert any("不一致" in w and "PO-100" in w for w in result.warnings)
    # 輸出以第一筆為準
    assert "台北市" in _output_cell(result, 2, 12)


# ------------------------------------------------------------------ 備忘錄來源


def test_memo_source_picks_line_or_main_column():
    """來源同時有「備忘錄」與「備忘錄 (主要)」時，依 memo_source 選擇來源欄。"""
    header = _E2E_HEADER + ["備忘錄", "備忘錄 (主要)"]
    row = _e2e_row("2027-05-01", 10) + ["明細行備註", "主要備註"]
    result_line = shipping_convert_rows(
        [header, row], MODE_ORDER, TEMPLATE_PATH, memo_source=MEMO_SOURCE_LINE
    )
    result_main = shipping_convert_rows(
        [header, row], MODE_ORDER, TEMPLATE_PATH, memo_source=MEMO_SOURCE_MAIN
    )
    assert "明細行備註" in _output_cell(result_line, 2, 19)
    assert "主要備註" in _output_cell(result_main, 2, 19)
    # 預設值維持原本行為（吃明細行「備忘錄」）
    result_default = shipping_convert_rows([header, row], MODE_ORDER, TEMPLATE_PATH)
    assert "明細行備註" in _output_cell(result_default, 2, 19)


def test_memo_line_equal_to_product_name_is_noise():
    """2026-08 新版報表明細行備忘錄=品名（NetSuite 回填），要過濾掉不進備註。"""
    header = _E2E_HEADER + ["備忘錄", "備忘錄 (主要)"]
    row = _e2e_row("2027-05-01", 10) + ["品名A", "真正的備註"]  # 備忘錄 = 項目名稱
    result = shipping_convert_rows(
        [header, row], MODE_ORDER, TEMPLATE_PATH, memo_source=MEMO_SOURCE_LINE
    )
    note = _output_cell(result, 2, 19)
    assert "品名A" not in note
    assert note.endswith("|")  # 溫馨提醒段為空
    assert any("與品名相同" in w for w in result.warnings)
    # 主要備忘錄不受影響
    result_main = shipping_convert_rows(
        [header, row], MODE_ORDER, TEMPLATE_PATH, memo_source=MEMO_SOURCE_MAIN
    )
    assert "真正的備註" in _output_cell(result_main, 2, 19)
    assert not any("與品名相同" in w for w in result_main.warnings)


def test_memo_source_falls_back_with_warning():
    """選擇的備忘錄欄不存在時，退回另一欄並提醒。"""
    header = _E2E_HEADER + ["備忘錄 (主要)"]
    row = _e2e_row("2027-05-01", 10) + ["主要備註"]
    result = shipping_convert_rows(
        [header, row], MODE_ORDER, TEMPLATE_PATH, memo_source=MEMO_SOURCE_LINE
    )
    assert "主要備註" in _output_cell(result, 2, 19)
    assert any("備忘錄" in w and "改用" in w for w in result.warnings)

    m00_result = m00_convert_rows(
        [header, row], MODE_ORDER, M00_TEMPLATE_PATH, memo_source=MEMO_SOURCE_LINE
    )
    assert any("改用" in w for w in m00_result.warnings)


# ------------------------------------------------------------------ 虛擬倉（輸出第 30 欄）


def _output_cell(result, row: int, col: int):
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(result.output_bytes))
    return wb["工作表1"].cell(row, col).value


def test_e2e_warehouse_from_source_location_column():
    """來源有地點/倉別欄時，逐列帶出 G 開頭代碼，不再固定填範本值。"""
    header = _E2E_HEADER + ["地點"]
    r1 = _e2e_row("2027-05-01", 10) + ["G30 出貨倉"]
    r2 = _e2e_row("2027-06-01", 5) + ["良品倉"]  # 無法辨識 → 範本預設 + 警告
    result = shipping_convert_rows([header, r1, r2], MODE_ORDER, TEMPLATE_PATH)
    assert _output_cell(result, 2, 30) == "G30"
    assert _output_cell(result, 3, 30) == "G10"
    assert any("無法辨識" in w for w in result.warnings)


def test_e2e_warehouse_display_material_falls_back_to_g90():
    """來源沒有倉別欄時，料號 9 開頭（陳列/宣傳品）輸出 G90 並附警告。"""
    normal = _e2e_row("2027-05-01", 10)
    display = _e2e_row("2027-05-01", 1)
    display[4] = "902126060002_X"   # 項目
    display[6] = "902126060002"     # DR_料號
    result = shipping_convert_rows([_E2E_HEADER, normal, display], MODE_ORDER, TEMPLATE_PATH)
    assert _output_cell(result, 2, 30) == "G10"   # 一般料號沿用範本預設
    assert _output_cell(result, 3, 30) == "G90"
    assert any("902126060002" in w and "G90" in w for w in result.warnings)


def test_e2e_warehouse_reads_netsuite_transfer_source_location_columns():
    """調撥單直接抓取:虛擬倉要吃 saved search 的「來源倉別」,不能落回 G10。

    欄名取自正式區實際回傳(調撥單未出貨明細 searchId=159)。過去
    _WAREHOUSE_ALIASES 只認「倉別」「地點」,「來源倉別」「來源地點」都對不上,
    整批默默填成範本預設值 G10,跟 NetSuite 上選的來源倉不一致。
    「目標倉別/目標地點」是收貨端,不可以被拿來當虛擬倉。
    """
    header = [
        "內部 ID", "文件編號", "日期", "DR_料號", "顯示名稱", "交易序號/批號",
        "來源倉別", "目標倉別", "出貨數量", "倉儲", "倉儲聯繫人", "倉儲電話",
        "倉儲地址", "備忘錄 (主要)", "項目", "來源地點", "目標地點",
        "DR_預計到貨日期", "DR_預計到貨時段",
    ]

    def row(material, qty, source_wh, source_loc):
        return [
            "1", "TO-1", "2026-08-01", material, "品名A", "2027-05-01",
            source_wh, "G10 新竹倉", qty, "門市A", "聯絡人", "0912345678",
            "10041台北市中正區路1號", "", material + "_X", source_loc,
            "客戶A", "2026-08-05", "早(9-12)",
        ]

    rows = [
        header,
        row("SKU001", 10, "G30 台北倉", "G30_台北倉"),
        # 「來源倉別」空白時要退到「來源地點」,不是直接放棄改用範本值
        row("SKU002", 5, "", "G80_桃園倉"),
    ]
    result = shipping_convert_rows(rows, MODE_TRANSFER, TEMPLATE_PATH)
    assert _output_cell(result, 2, 30) == "G30"
    assert _output_cell(result, 3, 30) == "G80"
    assert not any("沒有倉別欄" in w for w in result.warnings)


def test_e2e_warns_when_warehouse_cell_is_blank():
    """有倉別欄、但該列空白時也要警告一次(同樣會靜默變成 G10)。"""
    header = _E2E_HEADER + ["倉別"]
    filled = _e2e_row("2027-05-01", 10) + ["G30 出貨倉"]
    blank = _e2e_row("2027-06-01", 5) + [""]
    result = shipping_convert_rows([header, filled, blank], MODE_ORDER, TEMPLATE_PATH)
    assert _output_cell(result, 2, 30) == "G30"
    assert _output_cell(result, 3, 30) == "G10"
    blank_warnings = [w for w in result.warnings if "倉別欄是空白" in w]
    assert len(blank_warnings) == 1   # 同一則訊息不重複洗版
    assert not any("沒有倉別欄" in w for w in result.warnings)


def test_e2e_warns_when_source_has_no_warehouse_column():
    """整份報表都沒有倉別欄時要警告一次,否則虛擬倉整批默默變成 G10。"""
    result = shipping_convert_rows(
        [_E2E_HEADER, _e2e_row("2027-05-01", 10)], MODE_ORDER, TEMPLATE_PATH
    )
    assert _output_cell(result, 2, 30) == "G10"
    assert any("沒有倉別欄" in w and "G10" in w for w in result.warnings)


# ------------------------------------------------------------------ ED 訂單拆解

_SPLIT_HEADER = ["單號", "商品貨號", "商品名稱", "商品類型", "數量", "結帳單價", "含稅金額"]


def _split_map():
    """兩品套件（比率 0.65/0.35）＋三品套件（0.333/0.333/0.334，每套 2 個）。"""
    return {
        "29900001": [
            bundle_split.BundleComponent("A001", "單品A", "SPA", 1.0, 0.65),
            bundle_split.BundleComponent("A002", "單品B", "SPB", 1.0, 0.35),
        ],
        "29900002": [
            bundle_split.BundleComponent("B001", "面膜1", "", 2.0, 0.333),
            bundle_split.BundleComponent("B002", "面膜2", "", 2.0, 0.333),
            bundle_split.BundleComponent("B003", "面膜3", "", 2.0, 0.334),
        ],
    }


def _split_result(rows):
    return bundle_split.split_rows([_SPLIT_HEADER] + rows, _split_map(), "來源.xls", "對照.xlsx")


def test_split_expands_qty_and_keeps_amount_total():
    result = _split_result([
        ["S1", "29900001", "兩品組", "商品", 3, 500, 1000],
        ["S1", "101000000001", "一般單品", "商品", 2, 300, 600],
    ])
    assert result.output_rows == 3          # 套件拆 2 列 + 原樣保留 1 列
    assert result.split_rows == 1
    assert result.component_rows == 2
    assert result.amount_matched
    assert result.amount_before == result.amount_after == 1600
    qty_index = _SPLIT_HEADER.index("數量")
    sku_index = _SPLIT_HEADER.index("商品貨號")
    amount_index = _SPLIT_HEADER.index("含稅金額")
    exploded = [r for r in result.preview_rows if r[sku_index] in ("A001", "A002")]
    assert [r[qty_index] for r in exploded] == [3, 3]
    # 650/350，且來源金額是整數時拆出來也維持整數
    assert [r[amount_index] for r in exploded] == [650, 350]
    assert all(isinstance(r[amount_index], int) for r in exploded)


def test_split_rounding_remainder_goes_to_largest_ratio():
    # 864 × (0.333, 0.333, 0.334) 個別四捨五入會多 1 元，尾差要吃回比率最大者
    result = _split_result([["S1", "29900002", "三品組", "商品", 1, 999, 864]])
    amount_index = _SPLIT_HEADER.index("含稅金額")
    qty_index = _SPLIT_HEADER.index("數量")
    amounts = [r[amount_index] for r in result.preview_rows]
    assert sum(amounts) == 864
    assert amounts == [288, 288, 288]
    assert [r[qty_index] for r in result.preview_rows] == [2, 2, 2]  # 每套數量 2


def test_split_marks_and_labels_rows():
    result = _split_result([
        ["S1", "29900001", "兩品組", "商品", 1, 500, 1000],
        ["S1", "101000000001", "一般單品", "商品", 1, 300, 300],
    ])
    columns = result.preview_columns
    status = columns.index("拆解狀態")
    source_sku = columns.index("組合來源料號")
    assert result.preview_rows[0][status] == bundle_split.STATUS_SPLIT
    assert result.preview_rows[0][source_sku] == "29900001"
    assert result.preview_rows[0][columns.index("組合來源品名")] == "兩品組"
    assert result.preview_rows[-1][status] == bundle_split.STATUS_KEPT
    assert result.preview_rows[-1][source_sku] == ""


def test_split_gift_zero_amount_and_unmapped_bundle_warning():
    result = _split_result([
        ["S1", "29900001", "兩品組", "贈品", 1, 0, 0],
        ["S1", "29900099", "未對照的架上組", "贈品", 1, 0, 0],
    ])
    amount_index = _SPLIT_HEADER.index("含稅金額")
    assert [r[amount_index] for r in result.preview_rows[:2]] == [0, 0]
    assert [r["商品貨號"] for r in result.unmapped_rows] == ["29900099"]
    assert result.unmapped_rows[0]["出現列數"] == 1
    assert any("不在對照表" in w for w in result.warnings)


def test_split_298_is_a_real_product_not_an_unmapped_bundle():
    """298 開頭是正常組合品（ERP 有自己的品號與庫存），不拆也不該被當成漏對照。"""
    result = _split_result([["S1", "29826012", "杏仁酸煥膚透亮組", "商品", 1, 556, 556]])
    columns = result.preview_columns
    assert result.split_rows == 0
    assert result.preview_rows[0][columns.index("拆解狀態")] == bundle_split.STATUS_KEPT
    assert result.unmapped_rows == []
    assert result.warnings == []


def test_split_missing_required_column_raises():
    try:
        bundle_split.split_rows([["單號", "商品名稱"], ["S1", "X"]], _split_map())
    except bundle_split.BundleSplitError as exc:
        assert "商品貨號" in str(exc)
    else:
        raise AssertionError("缺少商品貨號欄位應該要報錯")


def test_load_bundle_map_normalizes_and_rejects_unknown_format():
    header = ["套件品號", "單品品號", "單品品名", "規格", "每套數量", "分攤比率"]
    headers = build_header_map(header)
    # 比率未正規化（加總 4）→ 載入後各 0.5
    rows = [["29900003", "C001", "單品C", "", 1, 2], ["29900003", "C002", "單品D", "", 1, 2]]
    comps = bundle_split._load_simple(rows, headers)["29900003"]
    assert [round(c.ratio, 6) for c in comps] == [0.5, 0.5]
    # 比率全空時平均分攤
    blank = [["29900004", "E001", "", "", 1, ""], ["29900004", "E002", "", "", 1, ""]]
    comps = bundle_split._load_simple(blank, headers)["29900004"]
    assert [round(c.ratio, 6) for c in comps] == [0.5, 0.5]
    # 欄位認不出來要報錯，而不是安靜回傳空對照表
    unknown = _spreadsheetml([["欄一", "欄二"], ["a", "b"]])
    try:
        bundle_split.load_bundle_map(unknown, "unknown.xls")
    except bundle_split.BundleSplitError as exc:
        assert "無法辨識" in str(exc)
    else:
        raise AssertionError("無法辨識的對照表格式應該要報錯")


def _spreadsheetml(rows: list[list[str]]) -> bytes:
    """組一份最小的 SpreadsheetML（NetSuite .xls 的格式），供讀檔路徑測試。"""
    cells = "".join(
        "<ss:Row>" + "".join(
            f'<ss:Cell><ss:Data ss:Type="String">{value}</ss:Data></ss:Cell>' for value in row
        ) + "</ss:Row>"
        for row in rows
    )
    return (
        '<?xml version="1.0"?>'
        '<Workbook xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">'
        f'<ss:Worksheet ss:Name="S1"><ss:Table>{cells}</ss:Table></ss:Worksheet></Workbook>'
    ).encode("utf-8")


def _read_sheet_rows(output_bytes: bytes, sheet_name: str) -> list[list]:
    """讀輸出檔某個分頁的所有列（含欄位列），供欄位/格式的回歸檢查。"""
    import io as _io

    import openpyxl

    wb = openpyxl.load_workbook(_io.BytesIO(output_bytes))
    return [list(row) for row in wb[sheet_name].iter_rows(values_only=True)]


# ------------------------------------------------------------------ table_filter

_FILTER_ROWS = [
    ["文件編號", "出貨客戶", "DR_預計到貨日期", "出貨數量", "備忘錄"],
    ["SO001", "屈臣氏", "2026-08-14", 12, "急件"],
    ["SO002", "康是美", "2026-08-15", 3, ""],
    ["SO003", "屈臣氏", "2026-08-20", 100, "補貨"],
    ["SO004", "寶雅", "", 7, "急件"],
]


def test_filter_keyword_is_and_across_columns():
    # 兩個詞要同時命中（可以分散在不同欄）
    assert table_filter.filter_indices(_FILTER_ROWS, "屈臣氏 補貨") == [2]
    assert table_filter.filter_indices(_FILTER_ROWS, "so00") == [0, 1, 2, 3]
    assert table_filter.filter_indices(_FILTER_ROWS, "不存在") == []
    # 沒給條件＝全部通過
    assert table_filter.filter_indices(_FILTER_ROWS) == [0, 1, 2, 3]


def test_filter_values_and_blank_option():
    values = table_filter.distinct_values(_FILTER_ROWS, 1)
    assert values == ["寶雅", "康是美", "屈臣氏"] or set(values) == {"寶雅", "康是美", "屈臣氏"}
    # 到貨日那欄有空白格 → 清單開頭要有「（空白）」選項
    dates = table_filter.distinct_values(_FILTER_ROWS, 2)
    assert dates[0] == table_filter.BLANK_LABEL
    assert table_filter.filter_indices(_FILTER_ROWS, value_filters={1: {"屈臣氏"}}) == [0, 2]
    # 只選「（空白）」時，只留該欄空白的列
    assert table_filter.filter_indices(
        _FILTER_ROWS, value_filters={2: {table_filter.BLANK_LABEL}}
    ) == [3]


def test_filter_text_contains_ignores_case():
    assert table_filter.filter_indices(_FILTER_ROWS, text_filters={0: "so00"}) == [0, 1, 2, 3]
    assert table_filter.filter_indices(_FILTER_ROWS, text_filters={0: "SO003"}) == [2]


def test_filter_date_range_and_column_detection():
    from datetime import date

    assert table_filter.is_date_column(_FILTER_ROWS, 2)
    # 「出貨數量」是小整數，不能被當成 Excel 日期序號誤判成日期欄
    assert not table_filter.is_date_column(_FILTER_ROWS, 3)
    assert not table_filter.is_date_column(_FILTER_ROWS, 0)
    picked = table_filter.filter_indices(
        _FILTER_ROWS, date_ranges={2: (date(2026, 8, 15), date(2026, 8, 20))}
    )
    assert picked == [1, 2]
    # 只給起日；日期解析不出來的列（空白）一律排除
    assert table_filter.filter_indices(_FILTER_ROWS, date_ranges={2: (date(2026, 8, 20), None)}) == [2]


def test_filter_conditions_are_combined():
    from datetime import date

    assert table_filter.filter_indices(
        _FILTER_ROWS,
        keyword="急件",
        value_filters={1: {"屈臣氏"}},
        date_ranges={2: (date(2026, 8, 1), date(2026, 8, 31))},
    ) == [0]


# ------------------------------------------------------------------ export_log


def test_export_log_extract_keys_prefers_document_number():
    rows = [
        ["內部 ID", "文件編號", "項目"],
        ["123", "SO001", "A"],
        ["124", "SO002", "B"],
    ]
    assert export_log.extract_keys(rows) == ["SO001", "SO002"]
    # 沒有文件編號欄時退回內部 ID
    rows2 = [["內部 ID", "項目"], ["123", "A"]]
    assert export_log.extract_keys(rows2) == ["123"]
    # 兩個都沒有 → 每列空字串（不會炸掉，只是標記不了）
    rows3 = [["項目", "數量"], ["A", 1]]
    assert export_log.extract_keys(rows3) == [""]


def test_export_log_mark_counts_and_dedupes():
    from datetime import datetime as dt

    log = export_log.empty_log()
    # 同一批裡重複出現的單號只算一次（一張單有多筆明細）
    export_log.mark(log, ["SO001", "SO001", "SO002", ""], "HCT 銷貨報表格式", dt(2026, 8, 14, 15, 3))
    assert set(log["records"]) == {"SO001", "SO002"}
    assert log["records"]["SO001"]["count"] == 1
    export_log.mark(log, ["SO001"], "M00 出貨格式", dt(2026, 8, 15, 9, 0))
    record = log["records"]["SO001"]
    assert record["count"] == 2
    assert record["first_at"] == "2026-08-14 15:03"
    assert record["last_at"] == "2026-08-15 09:00"
    assert record["formats"] == ["HCT 銷貨報表格式", "M00 出貨格式"]
    assert "已轉出 2 次" in export_log.describe(record)
    assert "×2" in export_log.short_label(record)


def test_export_log_roundtrip_and_corrupt_file():
    import tempfile
    from datetime import datetime as dt

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sub" / "log.json"
        export_log.record_export(path, ["SO001"], "HCT 銷貨報表格式", dt(2026, 8, 14, 15, 3))
        reloaded = export_log.load(path)
        assert reloaded["records"]["SO001"]["count"] == 1
        # 檔案壞掉不該擋住轉換，只是紀錄視為空的
        path.write_text("{ not json", encoding="utf-8")
        assert export_log.load(path)["records"] == {}
        # 不存在的檔案同理
        assert export_log.load(Path(tmp) / "nope.json")["records"] == {}


def test_export_log_merge_backup():
    base = export_log.empty_log()
    export_log.mark(base, ["SO001"], "HCT 銷貨報表格式", __import__("datetime").datetime(2026, 8, 14, 10, 0))
    other = export_log.empty_log()
    export_log.mark(other, ["SO001", "SO009"], "M00 出貨格式", __import__("datetime").datetime(2026, 8, 16, 10, 0))
    merged = export_log.merge(base, other)
    assert merged["records"]["SO001"]["count"] == 2
    assert merged["records"]["SO001"]["last_at"] == "2026-08-16 10:00"
    assert merged["records"]["SO001"]["first_at"] == "2026-08-14 10:00"
    assert "SO009" in merged["records"]


def test_export_log_loads_rejects_bad_payload():
    assert export_log.loads(b'{"version": 1, "records": {}}')["records"] == {}
    for bad in (b"[]", b"nope", b'{"records": 5}'):
        try:
            export_log.loads(bad)
        except ValueError:
            pass
        else:
            raise AssertionError(f"應該要拒絕：{bad!r}")


# ------------------------------------------------------------------ export_store（Google 試算表）


def test_sheet_rows_roundtrip():
    from datetime import datetime as dt

    log = export_log.empty_log()
    export_log.mark(log, ["SO001"], "HCT 銷貨報表格式", dt(2026, 8, 14, 15, 3))
    export_log.mark(log, ["SO001"], "M00 出貨格式", dt(2026, 8, 15, 9, 0))
    export_log.mark(log, ["SO002"], "HCT 銷貨報表格式", dt(2026, 8, 16, 8, 0))

    rows = export_store.log_to_rows(log)
    assert rows[0] == export_store.SHEET_HEADER
    # 最後轉出新的排前面
    assert [row[0] for row in rows[1:]] == ["SO002", "SO001"]

    back = export_store.rows_to_log(rows)
    assert back["records"]["SO001"]["count"] == 2
    assert back["records"]["SO001"]["first_at"] == "2026-08-14 15:03"
    assert back["records"]["SO001"]["last_at"] == "2026-08-15 09:00"
    assert back["records"]["SO001"]["formats"] == ["HCT 銷貨報表格式", "M00 出貨格式"]


def test_sheet_rows_tolerates_manual_edits():
    # 使用者手動調欄位順序、少幾欄、夾雜空白列，都不該讓整份紀錄壞掉
    rows = [
        ["最後轉出", "單號", "轉出次數"],
        ["2026-08-14 15:03", "SO001", "3"],
        ["", "", ""],
        ["2026-08-15 09:00", "SO002", "x"],  # 次數壞掉 → 至少算 1 次
    ]
    log = export_store.rows_to_log(rows)
    assert set(log["records"]) == {"SO001", "SO002"}
    assert log["records"]["SO001"]["count"] == 3
    assert log["records"]["SO002"]["count"] == 1
    # 空試算表
    assert export_store.rows_to_log([])["records"] == {}


def test_local_store_record_and_clear():
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        store = export_store.LocalStore(Path(tmp) / "log.json")
        assert store.is_empty()
        store.record(["SO001", "SO001"], "HCT 銷貨報表格式")
        assert list(store.load()["records"]) == ["SO001"]
        store.clear()
        assert store.is_empty()


def test_sheet_store_record_merges_into_sheet_content():
    """SheetStore.record 要先重讀試算表再合併，不能把別人剛寫的蓋掉。"""

    class FakeClient:
        spreadsheet_id = "fake"

        def __init__(self):
            self.pages = {}

        def read(self, worksheet, a1="A:Z"):
            return self.pages.get(worksheet, [])

        def write(self, worksheet, values, a1="A:Z"):
            self.pages[worksheet] = values

    client = FakeClient()
    store = export_store.SheetStore(client, "轉出紀錄")
    store.record(["SO001"], "HCT 銷貨報表格式")
    # 模擬別人在這期間寫進去的另一張單
    other = export_store.rows_to_log(client.read("轉出紀錄"))
    export_log.mark(other, ["SO999"], "M00 出貨格式")
    client.write("轉出紀錄", export_store.log_to_rows(other))

    store.record(["SO001"], "M00 出貨格式")
    final = store.load()["records"]
    assert set(final) == {"SO001", "SO999"}
    assert final["SO001"]["count"] == 2


def test_build_sheet_store_validates_config():
    # 有些貼上方式會讓 private_key 變成「字面上的 \n」而不是真的換行
    literal_nl = chr(92) + "n"
    good = {
        "spreadsheet_id": "https://docs.google.com/spreadsheets/d/ABC123/edit#gid=0",
        "service_account": {
            "client_email": "a@b.iam.gserviceaccount.com",
            "private_key": (
                "-----BEGIN PRIVATE KEY-----" + literal_nl + "x" + literal_nl + "-----END PRIVATE KEY-----"
            ),
            "token_uri": "https://oauth2.googleapis.com/token",
        },
    }
    store = export_store.build_sheet_store(good)
    assert store.client.spreadsheet_id == "ABC123"
    assert store.worksheet == export_store.DEFAULT_WORKSHEET
    # 字面上的 \n 要被救回成真正的換行，不然 google-auth 解不出金鑰
    assert chr(10) in store.client.info["private_key"]
    assert literal_nl not in store.client.info["private_key"]

    bad_configs = (
        {},
        {"spreadsheet_id": "x"},
        {"spreadsheet_id": "x", "service_account": {"client_email": "a"}},
    )
    for bad in bad_configs:
        try:
            export_store.build_sheet_store(bad)
        except gsheet.SheetError:
            pass
        else:
            raise AssertionError(f"應該要拒絕：{bad}")


def test_extract_spreadsheet_id():
    assert gsheet.extract_spreadsheet_id("ABC123") == "ABC123"
    assert gsheet.extract_spreadsheet_id(
        "https://docs.google.com/spreadsheets/d/1a-B_c/edit?gid=0#gid=0"
    ) == "1a-B_c"
    for bad in ("", "   ", "https://example.com/nope"):
        try:
            gsheet.extract_spreadsheet_id(bad)
        except gsheet.SheetError:
            pass
        else:
            raise AssertionError(f"應該要拒絕：{bad!r}")


# ------------------------------------------------------------------ 執行器


def main() -> int:
    failures = 0
    tests = [
        (name, fn) for name, fn in sorted(globals().items())
        if name.startswith("test_") and callable(fn)
    ]
    for name, fn in tests:
        try:
            fn()
        except AssertionError:
            import traceback
            failures += 1
            print(f"FAIL {name}")
            traceback.print_exc()
        else:
            print(f"ok   {name}")
    print(f"\n{len(tests) - failures}/{len(tests)} passed")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
