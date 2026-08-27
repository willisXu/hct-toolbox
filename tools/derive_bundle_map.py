# -*- coding: utf-8 -*-
"""從 NS 品項主檔的「規格」公式推導 299 組合對照表(補 ERP 匯出檔的缺漏)。

## 為什麼需要這支

ED 訂單拆解的內建對照表 `mappings/組合對照表.xlsx` 本來只能靠 ERP 匯出的
`bundles.csv`(鼎新 MDxxx 格式)維護,但那份檔案涵蓋不全 —— 2026-08-26 那份
只有 116 個套件、資料停在 2026/6/10、客戶代號只涵蓋 8 組,跳號嚴重
(…29926082、29926083 之後直接跳到 29926095)。

而 **NS 品項主檔的「規格」欄本身就是組合公式**:

    29926085 玻尿酸保濕精華液 團購組   規格  HHS030A*4+HHS015A
    29925164 玻尿酸保濕精華5入囤貨組   規格  HHS030A*5
    29925048 高效補水修護組            規格  HHT150A+DCS015A+HHL050A

規格代碼再用同一份主檔反查單品品號,就能把拆法還原出來。2026-08-26 用這個
方法把對照表從 119 補到 255 個套件。

## 收錄條件(三個都要成立才收,任一不成立就跳過留給人工)

  ① 「規格」是看得懂的組合公式(`A+B*2` 這種)
  ② 公式裡每個規格代碼在主檔**只對到一個單品**(非 29 開頭的品號)
  ③ 該單品在 ERP 匯出檔查得到標準售價(算分攤比率要用)

分攤比率 = 標準售價 × 數量加權,取兩位小數、尾差併入比率最大者(加總=1)。

## 準確度(2026-08-26 回測)

拿當時現行的 155 個套件用同一套規則重算:推導得出結果的 93 個裡 91 個組成與
數量完全相同(97.8%)。兩個不同的都是規格代碼 `AAS030A` 造成 —— 詳見
handoff.md,那是「公式沒跟上換版」不是規則錯。

## 用法

    python tools/derive_bundle_map.py <品項主檔.xlsx> <bundles.csv> [選項]

    --apply            實際寫回 mappings/組合對照表.xlsx(不帶只印出會新增什麼)
    --prefix 29925,29926   只處理這些前綴(預設值,即 2025/2026 年開的架上組)
    --prefix 299       不限年份 —— 主檔裡 2014~2024 的舊架上組還有兩百多個
                       推得出來,但那些多半早就不賣了,別無故灌進對照表

299 後面兩碼是年份:29925xxx = 2025 年開的、29926xxx = 2026 年。

寫回時所有儲存格一律文字格式,避免品號被 Excel 轉成數字/科學記號。
既有套件一律不動,只新增沒收錄過的。
"""
from __future__ import annotations

import csv
import os
import re
import sys
from collections import defaultdict

import openpyxl

# Windows 主控台預設 cp950,中文品名會變亂碼
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MAP_PATH = os.path.join(BASE, "mappings", "組合對照表.xlsx")

# 只處理架上組(299 開頭);單品品號排除 29 開頭,避免組合品被當成單品。
# 預設只碰 2025/2026 年開的(299 後兩碼是年份),舊年份要另外用 --prefix 指定。
DEFAULT_PREFIXES = ("29925", "29926")
TERM = re.compile(r"^([A-Za-z0-9]+)(?:\*(\d+))?$")


def load_item_master(path: str) -> tuple[list[dict], dict, dict]:
    """讀品項主檔 → (全部列, {品號: 列}, {規格代碼: [單品品號]})。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    it = wb.worksheets[0].iter_rows(values_only=True)
    hdr = [str(h).strip() if h else "" for h in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    for need in ("商品貨號", "規格"):
        if need not in ix:
            raise SystemExit(f"品項主檔缺「{need}」欄:{path}")
    items = [{h: ("" if r[i] is None else str(r[i]).strip())
              for h, i in ix.items() if i < len(r)} for r in it if r and r[0]]
    wb.close()
    by_code = {i["商品貨號"]: i for i in items}
    spec_to: dict[str, list[str]] = defaultdict(list)
    for i in items:
        if not i["商品貨號"].startswith("29") and i["規格"]:
            spec_to[i["規格"]].append(i["商品貨號"])
    return items, by_code, spec_to


def load_prices(path: str) -> dict[str, float]:
    """從 ERP 匯出檔取 {單品品號: 標準售價}(MD006 / MD016)。"""
    price: dict[str, float] = {}
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            sku = str(r.get("MD006_品號", "")).strip()
            try:
                p = float(r.get("MD016_標準售價") or 0)
            except (TypeError, ValueError):
                continue
            if sku and p > 0:
                price.setdefault(sku, p)
    return price


def parse_formula(spec: str) -> list[tuple[str, int]] | None:
    """「HHS030A*4+HHS015A」→ [("HHS030A", 4), ("HHS015A", 1)]。看不懂回 None。

    同一個規格代碼重複出現(RRT150A+RRT150A)就把數量加總成一項。
    """
    if not spec or ("+" not in spec and "*" not in spec):
        return None
    merged: dict[str, int] = defaultdict(int)
    order: list[str] = []
    for term in spec.split("+"):
        m = TERM.match(term.strip())
        if not m:
            return None       # 帶括號的版本記號(AAS030A(2507))等,不猜
        code, qty = m.group(1), int(m.group(2) or 1)
        if code not in merged:
            order.append(code)
        merged[code] += qty
    return [(c, merged[c]) for c in order]


def build_rows(bundle: str, by_code: dict, spec_to: dict,
               price: dict) -> tuple[list[list[str]] | None, str]:
    """一個套件 → 對照表列(套件品號/單品品號/單品品名/規格/每套數量/分攤比率)。"""
    item = by_code.get(bundle)
    if not item:
        return None, "主檔查無此品號"
    terms = parse_formula(item["規格"])
    if not terms:
        return None, "規格不是組合公式"
    rows, weights = [], []
    for code, qty in terms:
        cands = spec_to.get(code) or []
        if len(cands) != 1:
            return None, f"規格代碼對不到唯一單品({code} 有 {len(cands)} 個)"
        sku = cands[0]
        if sku not in price:
            return None, f"缺標準售價({sku})"
        weights.append(price[sku] * qty)
        rows.append([bundle, sku, by_code[sku].get("ERP品名(對照表)", ""),
                     code, str(qty), ""])
    total = sum(weights)
    if total <= 0:
        return None, "售價全為 0"
    ratios = [round(w / total, 2) for w in weights]
    diff = round(1 - sum(ratios), 2)
    if diff:      # 兩位小數的尾差併入比率最大者,加總才會剛好是 1
        ratios[max(range(len(ratios)), key=lambda k: weights[k])] += diff
    for row, r in zip(rows, ratios):
        row[5] = f"{round(r, 2):g}"
    return rows, ""


def read_map(path: str) -> tuple[str, list[str], list[list[str]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(it)]
    rows = [[("" if c is None else str(c).strip()) for c in r]
            for r in it if any(c not in (None, "") for c in r)]
    title = ws.title
    wb.close()
    return title, header, rows


def write_map(path: str, title: str, header: list[str], rows: list[list[str]]) -> None:
    """全表以文字格式寫出:品號之後用 Excel 開來編輯也不會被轉成數字。"""
    out = openpyxl.Workbook()
    sh = out.active
    sh.title = title
    for j, h in enumerate(header, 1):
        c = sh.cell(row=1, column=j, value=h)
        c.number_format = "@"
    for i, r in enumerate(rows, 2):
        for j in range(1, len(header) + 1):
            c = sh.cell(row=i, column=j, value=r[j - 1] if j - 1 < len(r) else "")
            c.number_format = "@"
    out.save(path)


def main(argv: list[str]) -> int:
    apply_ = "--apply" in argv
    prefixes = DEFAULT_PREFIXES
    argv = list(argv)
    if "--prefix" in argv:
        k = argv.index("--prefix")
        if k + 1 >= len(argv):
            print("--prefix 後面要接前綴,例如 --prefix 29925,29926")
            return 2
        prefixes = tuple(p.strip() for p in argv[k + 1].split(",") if p.strip())
        del argv[k:k + 2]
    args = [a for a in argv[1:] if not a.startswith("--")]
    if len(args) != 2:
        print(__doc__)
        return 2
    items, by_code, spec_to = load_item_master(args[0])
    price = load_prices(args[1])
    title, header, rows = read_map(MAP_PATH)
    have = {r[0] for r in rows}

    new_rows, blocked = [], defaultdict(list)
    for i in items:
        b = i["商品貨號"]
        if not b.startswith(prefixes) or b in have:
            continue
        built, err = build_rows(b, by_code, spec_to, price)
        if built:
            new_rows += built
        else:
            blocked[err.split("(")[0]].append(b)

    codes = sorted({r[0] for r in new_rows})
    print(f"對照表現有 {len(have)} 個套件、{len(rows)} 列(處理前綴:{'、'.join(prefixes)})")
    print(f"可新增:{len(codes)} 個套件、{len(new_rows)} 列")
    for k, v in sorted(blocked.items(), key=lambda kv: -len(kv[1])):
        print(f"  跳過 — {k}:{len(v)} 個  例:{v[:5]}")
    if not apply_:
        print("\n(未帶 --apply,沒有寫檔。上面是會新增的內容。)")
        for r in new_rows[:20]:
            print("   " + "  ".join(r))
        if len(new_rows) > 20:
            print(f"   …另有 {len(new_rows) - 20} 列")
        return 0

    rows += new_rows
    rows.sort(key=lambda r: r[0])
    write_map(MAP_PATH, title, header, rows)
    print(f"\n已寫回 {MAP_PATH}:{len(have) + len(codes)} 個套件、{len(rows)} 列")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
